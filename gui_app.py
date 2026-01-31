"""
电商工具箱 - GUI 应用
基于 PyQt5 的图形界面
"""

import sys
import os
import warnings
from typing import Dict

# 禁用 PyQt5 的 DeprecationWarning
warnings.filterwarnings('ignore', category=DeprecationWarning)

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTableWidget, QTableWidgetItem, QTabWidget,
    QLabel, QComboBox, QSpinBox, QDoubleSpinBox, QMessageBox, QProgressBar,
    QTextEdit, QSplitter, QHeaderView, QStatusBar, QDialog, QGridLayout,
    QSpacerItem, QSizePolicy, QStackedWidget, QLineEdit, QButtonGroup, QInputDialog, QCheckBox,
    QScrollArea, QGroupBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QColor, QIcon, QPixmap, QPainter
from main.common.config import ConfigManager
from main.label.label_generator import LabelGenerator
from main.label.carton_mark_generator import CartonMarkGenerator




class ExportThread(QThread):
    """导出线程"""
    finished = pyqtSignal()
    error = pyqtSignal(str)
    success = pyqtSignal(str)
    
    def __init__(self, gui_instance, export_folder, label_type, export_date, export_year, export_label, export_box, export_order):
        super().__init__()
        self.gui = gui_instance
        self.export_folder = export_folder
        self.label_type = label_type
        self.export_date = export_date
        self.export_year = export_year
        self.export_label = export_label
        self.export_box = export_box
        self.export_order = export_order
    
    def run(self):
        try:
            if self.export_label:
                self.gui.generate_labels_from_workbook(
                    self.export_folder, self.label_type, self.export_date, self.export_year
                )
            
            if self.export_box:
                self.gui.generate_carton_marks_from_workbook(
                    self.export_folder, self.label_type, self.export_date, self.export_year
                )
            
            if self.export_order:
                self.gui.generate_order_form(self.label_type, self.export_date, self.export_year)
            
            self.success.emit("导出完成")
        except Exception as e:
            self.error.emit(f"导出失败: {str(e)}")
        finally:
            self.finished.emit()


class PropellerFilenameDialog(QDialog):
    """螺旋桨文件名输入对话框 - 批量输入"""
    
    def __init__(self, parent, new_skus, existing_filenames):
        super().__init__(parent)
        self.setWindowTitle("输入螺旋桨文件名")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowType.WindowContextHelpButtonHint)
        self.setModal(True)
        self.setMinimumWidth(500)
        
        self.sku_inputs = {}  # 存储 SKU -> 输入框的映射
        self.existing_filenames = existing_filenames  # 已有的文件名集合
        
        layout = QVBoxLayout()
        
        # 标签
        label = QLabel("请为以下螺旋桨输入文件名：")
        layout.addWidget(label)
        
        # 创建滚动区域用于多个输入框
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        
        # 为每个新SKU创建一行（商品编号 + 输入框）
        for sku in new_skus:
            row_layout = QHBoxLayout()
            
            # 左侧：商品编号标签
            sku_label = QLabel(sku)
            sku_label.setMinimumWidth(150)
            row_layout.addWidget(sku_label)
            
            # 右侧：输入框
            text_input = QLineEdit()
            text_input.setText(sku)  # 默认值为SKU本身
            row_layout.addWidget(text_input)
            
            self.sku_inputs[sku] = text_input
            scroll_layout.addLayout(row_layout)
        
        scroll_widget.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)
        
        # 确定按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        ok_button = QPushButton("确定")
        ok_button.clicked.connect(self.validate_and_accept)
        button_layout.addWidget(ok_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def validate_and_accept(self):
        """验证输入，检查是否有重名"""
        filenames = {}
        empty_skus = []
        duplicates_with_existing = []
        duplicates_within_new = []
        
        for sku, input_widget in self.sku_inputs.items():
            filename = input_widget.text().strip()
            
            if not filename:
                empty_skus.append(sku)
                continue
            
            # 检查是否与已有的映射重名
            if filename in self.existing_filenames:
                duplicates_with_existing.append(f"{sku} -> {filename}")
            
            # 检查是否与新增的其他文件名重复
            if filename in filenames.values():
                duplicates_within_new.append(f"{sku} -> {filename}")
            
            filenames[sku] = filename
        
        # 处理为空的情况
        if empty_skus:
            msg = "以下商品编号的文件名为空：\n" + "\n".join(empty_skus) + "\n\n是否跳过这些商品编号？"
            reply = QMessageBox.question(self, "文件名为空", msg)
            if reply != QMessageBox.Yes:
                return
            # 从sku_inputs中移除为空的项
            for sku in empty_skus:
                del self.sku_inputs[sku]
        
        # 检查与已有映射的重名
        if duplicates_with_existing:
            msg = "以下文件名与已有映射重名，请重新填写：\n" + "\n".join(duplicates_with_existing)
            QMessageBox.warning(self, "重名警告", msg)
            return
        
        # 检查新增文件名之间的重复
        if duplicates_within_new:
            msg = "以下文件名在新增的商品编号中重复，请重新填写：\n" + "\n".join(duplicates_within_new)
            QMessageBox.warning(self, "重复警告", msg)
            return
        
        # 验证通过，接受对话框
        self.accept()
    
    def get_filenames(self):
        """返回 SKU -> 文件名的映射"""
        result = {}
        for sku, input_widget in self.sku_inputs.items():
            result[sku] = input_widget.text()
        return result


class KeywordAnalyzerGUI(QMainWindow):
    """关键词分析工具 GUI"""
    
    def __init__(self):
        super().__init__()
        self.config_manager = ConfigManager()
        
        self.setWindowTitle("ToolHub")
        
        # 设置固定窗口大小
        self.setFixedSize(900, 600)
        self.setGeometry(100, 100, 900, 600)
        
        # Excel导入相关
        self.current_excel_file = None
        self.current_excel_workbook = None
        
        # 品牌-工作表映射关系
        self.sheet_mapping = self.load_sheet_mapping()
        
        # 品牌-工作表规则关系
        self.sheet_rules = self.load_sheet_rules()
        
        # 记住上次打开的文件路径
        self.last_file_path = self.config_manager.get('last_csv_path', os.path.expanduser('~/Desktop'))
        
        # 应用版本和导出路径
        self.app_version = "1.1"
        self.export_path = self.config_manager.get('export_path', os.path.expanduser('~/Desktop'))
        
        # 检测更新
        self.check_for_updates()
        
        self.init_ui()
    
    def init_ui(self):
        """初始化 UI"""
        # 创建中央 widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局（水平）
        main_layout = QHBoxLayout()
        
        # 侧边栏
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setSpacing(10)
        
        # 工具按钮 - 使用互斥按钮组
        self.button_group = QButtonGroup()
        
        self.label_btn = QPushButton("标签箱唛")
        self.label_btn.setCheckable(True)
        self.label_btn.setChecked(True)
        self.label_btn.clicked.connect(lambda: self.switch_tool("label"))
        self.button_group.addButton(self.label_btn, 0)
        sidebar_layout.addWidget(self.label_btn)
        
        self.video_btn = QPushButton("种草视频")
        self.video_btn.setCheckable(True)
        self.video_btn.clicked.connect(lambda: self.switch_tool("video"))
        self.button_group.addButton(self.video_btn, 1)
        sidebar_layout.addWidget(self.video_btn)
        
        sidebar_layout.addStretch()
        
        # 设置按钮
        settings_btn = QPushButton("设置")
        settings_btn.clicked.connect(self.show_app_settings)
        sidebar_layout.addWidget(settings_btn)
        
        # 侧边栏容器
        sidebar_widget = QWidget()
        sidebar_widget.setLayout(sidebar_layout)
        sidebar_widget.setFixedWidth(100)
        
        main_layout.addWidget(sidebar_widget)
        
        # 工具容器（堆栈）
        self.tool_stack = QStackedWidget()
        
        # 标签箱唛工具页面
        label_widget = self.create_label_tool()
        self.tool_stack.addWidget(label_widget)
        
        # 种草助手页面
        video_widget = self.create_video_tool()
        self.tool_stack.addWidget(video_widget)
        
        main_layout.addWidget(self.tool_stack)
        
        central_widget.setLayout(main_layout)
    
    def switch_tool(self, tool_name):
        """切换工具"""
        if tool_name == "label":
            self.tool_stack.setCurrentIndex(0)
        elif tool_name == "video":
            self.tool_stack.setCurrentIndex(1)
    
    def show_app_settings(self):
        """显示应用设置对话框"""
        dialog = AppSettingsDialog(self, self.app_version, self.export_path)
        dialog.exec_()
        # 对话框关闭时保存导出路径（无论是关闭还是确定）
        new_export_path = dialog.get_export_path()
        if new_export_path != self.export_path:
            self.export_path = new_export_path
            self.config_manager.set('export_path', self.export_path)
            self.config_manager.save_config()
    
    def check_for_updates(self):
        """检测更新（从GitHub）"""
        try:
            import urllib.request
            import json
            import ssl
            
            # GitHub API URL
            github_api_url = "https://api.github.com/repos/corbin-xu/ToolHub/releases/latest"
            
            print("[DEBUG] ========== 开始检测更新 ==========")
            print(f"[DEBUG] 当前版本: {self.app_version}")
            print(f"[DEBUG] API URL: {github_api_url}")
            
            try:
                print("[DEBUG] 正在创建 SSL 上下文...")
                # 创建 SSL 上下文，禁用证书验证（用于 PyInstaller 打包的应用）
                ssl_context = ssl.create_default_context()
                ssl_context.check_hostname = False
                ssl_context.verify_mode = ssl.CERT_NONE
                print("[DEBUG] SSL 上下文创建成功")
                
                print("[DEBUG] 正在创建请求...")
                request = urllib.request.Request(github_api_url)
                request.add_header('User-Agent', 'ToolHub/1.2')
                print("[DEBUG] 请求创建成功")
                
                print("[DEBUG] 正在连接到 GitHub API...")
                with urllib.request.urlopen(request, context=ssl_context, timeout=10) as response:
                    print(f"[DEBUG] 连接成功！状态码: {response.status}")
                    print("[DEBUG] 正在读取响应数据...")
                    response_data = response.read().decode()
                    print(f"[DEBUG] 响应数据长度: {len(response_data)} 字节")
                    
                    data = json.loads(response_data)
                    latest_version = data.get('tag_name', 'unknown').lstrip('v')
                    print(f"[DEBUG] 最新版本: {latest_version}")
                    
                    if latest_version > self.app_version:
                        print(f"[INFO] 发现新版本: {latest_version}，当前版本: {self.app_version}")
                    else:
                        print(f"[INFO] 已是最新版本")
                        
            except urllib.error.URLError as e:
                print(f"[ERROR] URLError 异常!")
                print(f"[ERROR] 错误信息: {e}")
                print(f"[ERROR] 错误原因: {e.reason if hasattr(e, 'reason') else '未知'}")
                import traceback
                traceback.print_exc()
                    
            except urllib.error.HTTPError as e:
                print(f"[ERROR] HTTPError 异常!")
                print(f"[ERROR] HTTP 状态码: {e.code}")
                print(f"[ERROR] 错误信息: {e.reason}")
                
            except json.JSONDecodeError as e:
                print(f"[ERROR] JSON 解析异常!")
                print(f"[ERROR] 错误信息: {e}")
                
            except Exception as e:
                print(f"[ERROR] 未知异常!")
                print(f"[ERROR] 异常类型: {type(e).__name__}")
                print(f"[ERROR] 错误信息: {str(e)}")
                import traceback
                traceback.print_exc()
                
            print("[DEBUG] ========== 检测更新完成 ==========")
        
        except Exception as e:
            print(f"[ERROR] 外层异常: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def on_sn_prefix_changed(self, prefix):
        """品牌前缀改变时的处理"""
        # 年份下拉栏始终显示 2025/2026
        # 在预览和生成时根据前缀转换为相应格式
        pass
    
    def import_excel_file(self):
        """导入Excel文件"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "警告", "需要安装openpyxl库。请运行: pip install openpyxl")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        try:
            # 获取文件名（不含路径和扩展名）
            import os
            file_name = os.path.basename(file_path)
            file_name_without_ext = os.path.splitext(file_name)[0]
            
            # 加载Excel文件，使用 data_only=True 显示公式的计算结果
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            
            # 只显示文件名
            preview_text = file_name
            
            # 从文件名中检测类型
            detected_type = ""
            if "3C" in file_name_without_ext or "3c" in file_name_without_ext:
                detected_type = "3C"
            elif "玩具" in file_name_without_ext or "toy" in file_name_without_ext.lower():
                detected_type = "玩具"
            elif "螺旋桨" in file_name_without_ext or "propeller" in file_name_without_ext.lower():
                detected_type = "螺旋桨"
            
            if detected_type:
                # 自动选中下拉栏
                self.preview_type_combo.setCurrentText(detected_type)
            else:
                # 无法自动识别，弹出对话框让用户选择
                items = ["3C", "玩具"]
                selected_type, ok = QInputDialog.getItem(
                    self, "选择类型", "无法自动识别文件类型，请选择:", items, 0, False
                )
                if ok and selected_type:
                    detected_type = selected_type
                    self.preview_type_combo.setCurrentText(detected_type)
                else:
                    return
            
            self.current_excel_file = file_path
            self.current_excel_workbook = workbook
            
            # 填充工作表选择器
            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheet_names)
            self.sheet_combo.blockSignals(False)
            
            # 默认选择第一个工作表
            if sheet_names:
                self.sheet_combo.setCurrentIndex(0)
                # 手动触发工作表改变事件以显示数据
                self.on_sheet_changed(sheet_names[0])
            
            print(f"[DEBUG] import_excel_file: detected_type={detected_type}")
            print(f"[DEBUG] sheet_mapping keys: {list(self.sheet_mapping.keys())}")
            print(f"[DEBUG] workbook.sheetnames: {workbook.sheetnames}")
            
            # 从表格中识别日期和年份
            detected_date, detected_year = self.extract_date_from_workbook(workbook, detected_type)
            print(f"[DEBUG] 识别结果: detected_date={detected_date}, detected_year={detected_year}")
            
            if detected_date:
                self.export_date_input.setText(detected_date)
            if detected_year:
                self.export_year_input.setText(detected_year)
            
            # 检查工作表是否匹配设置
            self.check_sheet_mapping(detected_type, workbook)
            
            # 更新文件名标签
            self.file_name_label.setText(f"已加载: {file_name}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入Excel文件失败: {str(e)}")
    
    def extract_date_from_workbook(self, workbook, label_type):
        """从工作表中识别日期和年份，检查所有映射工作表的日期是否一致"""
        try:
            print(f"[DEBUG] extract_date_from_workbook: label_type={label_type}")
            
            # 根据类型获取对应的工作表
            if label_type not in self.sheet_mapping:
                print(f"[DEBUG] label_type 不在 sheet_mapping 中")
                return "", ""
            
            brands = self.sheet_mapping[label_type]
            print(f"[DEBUG] brands={brands}")
            
            if not brands:
                print(f"[DEBUG] brands 为空")
                return "", ""
            
            # 收集所有工作表的日期和年份
            all_dates = []
            all_years = []
            
            # 获取识别配置
            recognize_config = self.config_manager.get('recognize_config', {
                'sheet_name': '勿改动！机型',
                'row': 2,
                'column': 10  # J列
            })
            
            sheet_name = recognize_config.get('sheet_name', '勿改动！机型')
            row = recognize_config.get('row', 2)
            column = recognize_config.get('column', 10)
            
            # 检查是否有指定的工作表
            if sheet_name in workbook.sheetnames:
                # 从指定工作表的指定单元格读取日期
                lookup_ws = workbook[sheet_name]
                date_cell = lookup_ws.cell(row, column)
                if date_cell.value:
                    date_value = str(date_cell.value).strip()
                    print(f"[DEBUG] 从{sheet_name}工作表{row}行{column}列读取日期: {date_value}")
                    
                    if label_type == "3C":
                        # 3C: YYMMDD 格式，如 251208
                        if len(date_value) >= 6 and date_value.isdigit():
                            year_part = date_value[:2]  # 前2位是年份
                            date_part = date_value[2:6]  # 中间4位是MMDD
                            all_years.append(f"20{year_part}")
                            all_dates.append(date_part)
                            print(f"[DEBUG] 3C格式: 年份={year_part}, 日期={date_part}")
                    elif label_type == "玩具":
                        # 玩具: YYYYMMDD 格式，如 20261208
                        if len(date_value) >= 8 and date_value.isdigit():
                            year_part = date_value[:4]  # 前4位是年份
                            date_part = date_value[4:8]  # 后4位是MMDD
                            all_years.append(year_part)
                            all_dates.append(date_part)
                            print(f"[DEBUG] 玩具格式: 年份={year_part}, 日期={date_part}")
            
            # 如果没有找到日期，尝试从配货表中查找
            if not all_dates:
                for brand, sheets in brands.items():
                    for sheet_name in sheets:
                        if sheet_name not in workbook.sheetnames:
                            print(f"[DEBUG] 工作表不存在: {sheet_name}")
                            continue
                        
                        ws = workbook[sheet_name]
                        
                        # 从第 2 行开始查找 D 列的 SN 码
                        for row in range(2, min(20, ws.max_row + 1)):
                            sn_cell = ws.cell(row, 4)  # D 列
                            if sn_cell.value:
                                sn_value = sn_cell.value
                                sn_str = str(sn_value).strip()
                                
                                print(f"[DEBUG] 工作表: {sheet_name}, 行: {row}, SN: {sn_str}")
                                
                                # 如果是公式，尝试解析VLOOKUP公式
                                if sn_str.startswith("="):
                                    print(f"[DEBUG] 检测到公式: {sn_str}")
                                    sn_str = self.evaluate_vlookup_formula(sn_str, workbook, ws)
                                    if not sn_str:
                                        print(f"[DEBUG] 无法解析公式")
                                        continue
                                
                                # 从 SN 中提取日期
                                date = self.extract_date_from_sn(sn_str)
                                # 从 SN 中提取年份
                                year = self.extract_year_from_sn(sn_str, label_type)
                                print(f"[DEBUG] 提取的日期: {date}, 年份: {year}")
                                
                                if date:
                                    all_dates.append(date)
                                if year:
                                    all_years.append(year)
                                
                                # 找到第一个有效的日期后就停止
                                if date:
                                    break
                        
                        # 如果已经找到日期，就停止查找其他工作表
                        if all_dates:
                            break
            
            if not all_dates or not all_years:
                print(f"[DEBUG] 未找到有效的日期或年份")
                return "", ""
            
            # 检查所有日期是否一致
            unique_dates = set(all_dates)
            unique_years = set(all_years)
            
            print(f"[DEBUG] 所有日期: {all_dates}, 唯一日期: {unique_dates}")
            print(f"[DEBUG] 所有年份: {all_years}, 唯一年份: {unique_years}")
            
            if len(unique_dates) > 1 or len(unique_years) > 1:
                # 日期或年份不一致，弹窗提示
                date_str = ", ".join(unique_dates)
                year_str = ", ".join(unique_years)
                QMessageBox.warning(
                    self, 
                    "警告", 
                    f"不同工作表中的日期或年份不一致:\n日期: {date_str}\n年份: {year_str}\n请手动修改"
                )
                return "", ""
            
            # 日期和年份一致，返回第一个
            return all_dates[0], all_years[0]
        
        except Exception as e:
            print(f"[DEBUG] 从工作表中识别日期失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return "", ""
    
    def evaluate_vlookup_formula(self, formula, workbook, current_sheet=None):
        """解析VLOOKUP公式并返回结果"""
        try:
            import re
            
            # 匹配 VLOOKUP 公式，支持列范围（如A:G）和单元格范围（如A1:G100）
            # 例如: =VLOOKUP(A2,'Sheet1'!A:G,7,0) 或 =VLOOKUP(A2,Sheet1!A1:G100,2,FALSE)
            pattern = r"=VLOOKUP\s*\(\s*([^,]+)\s*,\s*'?([^'!]+)'?\s*!\s*([A-Z]+(?::\s*[A-Z]+|\d+:[A-Z]+\d+))\s*,\s*(\d+)\s*,\s*([^)]+)\s*\)"
            match = re.match(pattern, formula, re.IGNORECASE)
            
            if not match:
                print(f"[DEBUG] 无法匹配VLOOKUP公式: {formula}")
                return ""
            
            lookup_value_ref = match.group(1).strip()
            lookup_sheet = match.group(2).strip()
            lookup_range = match.group(3).strip()
            col_index = int(match.group(4))
            
            print(f"[DEBUG] VLOOKUP参数: lookup_value_ref={lookup_value_ref}, lookup_sheet={lookup_sheet}, lookup_range={lookup_range}, col_index={col_index}")
            
            # 移除绝对引用符号
            lookup_value_ref = lookup_value_ref.replace('$', '')
            
            # 解析单元格引用（如 A2）
            cell_match = re.match(r'([A-Z]+)(\d+)', lookup_value_ref, re.IGNORECASE)
            if not cell_match:
                print(f"[DEBUG] 无法解析单元格引用: {lookup_value_ref}")
                return ""
            
            col_letter = cell_match.group(1)
            row_num = int(cell_match.group(2))
            
            # 获取当前工作表
            if current_sheet is None:
                current_sheet = workbook.active
            current_ws = current_sheet
            lookup_cell = current_ws[f"{col_letter}{row_num}"]
            lookup_value = lookup_cell.value
            
            print(f"[DEBUG] 查找值: {lookup_value}")
            
            # 获取查找表所在的工作表
            if lookup_sheet not in workbook.sheetnames:
                print(f"[DEBUG] 查找表工作表不存在: {lookup_sheet}")
                return ""
            
            lookup_ws = workbook[lookup_sheet]
            
            # 解析查找范围
            lookup_range_clean = lookup_range.replace('$', '').replace(' ', '')
            
            # 检查是否是列范围（如 A:G）
            if ':' in lookup_range_clean and not any(c.isdigit() for c in lookup_range_clean.split(':')[0]):
                # 列范围，如 A:G
                range_parts = lookup_range_clean.split(':')
                start_col = range_parts[0]
                end_col = range_parts[1]
                start_row = 1
                end_row = lookup_ws.max_row
            else:
                # 单元格范围，如 A1:G100
                range_match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', lookup_range_clean, re.IGNORECASE)
                if not range_match:
                    print(f"[DEBUG] 无法解析查找范围: {lookup_range}")
                    return ""
                
                start_col = range_match.group(1)
                start_row = int(range_match.group(2))
                end_col = range_match.group(3)
                end_row = int(range_match.group(4))
            
            # 转换列字母为数字
            start_col_num = ord(start_col.upper()) - ord('A') + 1
            
            print(f"[DEBUG] 查找范围: 行{start_row}-{end_row}, 列{start_col_num}")
            
            # 在查找表中查找值
            for row in range(start_row, end_row + 1):
                cell_value = lookup_ws.cell(row, start_col_num).value
                if cell_value == lookup_value:
                    # 找到匹配的行，返回指定列的值
                    result_cell = lookup_ws.cell(row, start_col_num + col_index - 1)
                    result = result_cell.value
                    print(f"[DEBUG] VLOOKUP结果: {result}")
                    return str(result).strip() if result else ""
            
            print(f"[DEBUG] 在查找表中未找到匹配值: {lookup_value}")
            return ""
        
        except Exception as e:
            print(f"[DEBUG] 解析VLOOKUP公式失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return ""
    def extract_date_from_sn(self, sn):
        """从 SN 中提取日期（MMDD 格式）"""
        try:
            sn_str = str(sn).strip()
            if not sn_str:
                return ""
            
            print(f"[DEBUG] extract_date_from_sn: sn_str={sn_str}, len={len(sn_str)}")
            
            # 日期是倒数第 4~7 位（从右往左数）
            if len(sn_str) >= 7:
                date = sn_str[-7:-3]  # 倒数第 7 到 4 位
                print(f"[DEBUG] 提取的日期: {date}")
                # 验证日期格式（MMDD）
                if date.isdigit() and len(date) == 4:
                    return date
            
            return ""
        except Exception as e:
            print(f"[DEBUG] 从 SN 中提取日期失败: {str(e)}")
            return ""
    
    def extract_year_from_sn(self, sn, label_type):
        """从 SN 中提取年份"""
        try:
            sn_str = str(sn).strip()
            if not sn_str:
                return ""
            
            print(f"[DEBUG] extract_year_from_sn: sn_str={sn_str}, label_type={label_type}, len={len(sn_str)}")
            
            # 根据类型确定年份位置
            if label_type == "3C":
                # 3C（SG）：年份是倒数第 8~9 位（2 位）
                if len(sn_str) >= 9:
                    year = sn_str[-9:-7]  # 倒数第 9 到 8 位
                    print(f"[DEBUG] 3C 提取的年份: {year}")
                    if year.isdigit() and len(year) == 2:
                        # 转换为完整年份（25 -> 2025）
                        return f"20{year}"
            elif label_type == "玩具":
                # 玩具（NB）：年份是倒数第 8~11 位（4 位）
                if len(sn_str) >= 11:
                    year = sn_str[-11:-7]  # 倒数第 11 到 8 位
                    print(f"[DEBUG] 玩具提取的年份: {year}")
                    if year.isdigit() and len(year) == 4:
                        return year
            
            return ""
        except Exception as e:
            print(f"[DEBUG] 从 SN 中提取年份失败: {str(e)}")
            return ""
    
    def parse_sn(self, sn):
        """
        从SN序列号中解析出各个部分
        
        SG格式: SG + 序号(1-3位) + 年份(2位) + 日期(4位) + 批次(3位)
        NB格式: NB + 序号(1-3位) + 年份(4位) + 日期(4位) + 批次(3位)
        
        Args:
            sn: 完整的SN序列号，如 "SG001202601001" 或 "NB001202501001"
            
        Returns:
            dict: 包含 prefix, seq, year, date, batch 的字典
        """
        if not sn or len(sn) < 2:
            return None
        
        prefix = sn[:2]
        rest = sn[2:]
        
        try:
            if prefix in ["SG", "SN"]:
                # SG/SN: 序号(1-3位) + 年份(2位) + 日期(4位) + 批次(3位)
                # 从后往前: 批次(3) + 日期(4) + 年份(2) + 序号(剩余)
                batch = rest[-3:]
                date = rest[-7:-3]
                year = rest[-9:-7]
                seq = rest[:-9]
                
                return {
                    'prefix': prefix,
                    'seq': seq,
                    'year': year,
                    'date': date,
                    'batch': batch
                }
            elif prefix == "NB":
                # NB: 序号(1-3位) + 年份(4位) + 日期(4位) + 批次(3位)
                # 从后往前: 批次(3) + 日期(4) + 年份(4) + 序号(剩余)
                batch = rest[-3:]
                date = rest[-7:-3]
                year = rest[-11:-7]
                seq = rest[:-11]
                
                return {
                    'prefix': prefix,
                    'seq': seq,
                    'year': year,
                    'date': date,
                    'batch': batch
                }
        except:
            return None
        
        return None
    
    def on_type_changed(self, type_name):
        """筛选类型改变时的处理"""
        # 根据类型自动设置 SN 前缀
        if type_name == "3C":
            self.label_sn_prefix.blockSignals(True)
            self.label_sn_prefix.setCurrentText("SG")
            self.label_sn_prefix.blockSignals(False)
            # 启用69码和SN字段
            self.label_code69_label.setEnabled(True)
            self.label_code69_input.setEnabled(True)
            self.label_sn_label.setEnabled(True)
            self.label_sn_prefix.setEnabled(True)
            self.label_sn_seq.setEnabled(True)
            self.label_sn_year.setEnabled(True)
            self.label_sn_year.setText("2026")  # 恢复默认值
            self.label_sn_date.setEnabled(True)
            self.label_sn_batch.setEnabled(True)
            self.label_sn_batch.setText("001")  # 恢复默认值
        elif type_name == "玩具":
            self.label_sn_prefix.blockSignals(True)
            self.label_sn_prefix.setCurrentText("NB")
            self.label_sn_prefix.blockSignals(False)
            # 启用69码和SN字段
            self.label_code69_label.setEnabled(True)
            self.label_code69_input.setEnabled(True)
            self.label_sn_label.setEnabled(True)
            self.label_sn_prefix.setEnabled(True)
            self.label_sn_seq.setEnabled(True)
            self.label_sn_year.setEnabled(True)
            self.label_sn_year.setText("2026")  # 恢复默认值
            self.label_sn_date.setEnabled(True)
            self.label_sn_batch.setEnabled(True)
            self.label_sn_batch.setText("001")  # 恢复默认值
        elif type_name == "螺旋桨":
            # 禁用69码和SN字段，但保持日期启用
            self.label_code69_label.setEnabled(False)
            self.label_code69_input.setEnabled(False)
            self.label_code69_input.clear()
            self.label_sn_label.setEnabled(False)
            self.label_sn_prefix.setEnabled(False)
            self.label_sn_prefix.setCurrentIndex(-1)  # 空选
            self.label_sn_seq.setEnabled(False)
            self.label_sn_seq.clear()
            self.label_sn_year.setEnabled(False)
            self.label_sn_year.clear()  # 清空填写框
            self.label_sn_date.setEnabled(True)  # 日期保持启用
            self.label_sn_batch.setEnabled(False)
            self.label_sn_batch.clear()
    
    def on_box_type_changed(self, type_name):
        """箱唛筛选类型改变时的处理"""
        # 根据类型自动填充商家名称
        if type_name == "3C":
            self.box_vendor_input.setText("ststthmyyx")
        elif type_name == "玩具":
            self.box_vendor_input.setText("stsnb")
    
    @staticmethod
    def convert_year_format(year_str: str, prefix: str) -> str:
        """
        根据 SN 前缀转换年份格式
        
        Args:
            year_str: 年份字符串 (2025 或 2026)
            prefix: SN 前缀 (SG, SN, NB)
            
        Returns:
            转换后的年份字符串
        """
        if prefix in ["SG", "SN"]:
            # SG/SN: 2位年份 (25, 26)
            if year_str == "2025":
                return "25"
            elif year_str == "2026":
                return "26"
        elif prefix == "NB":
            # NB: 4位年份 (2025, 2026)
            return year_str
        
        return year_str
    
    def on_brand_changed(self, brand):
        """品牌改变时的处理"""
        # 品牌改变时不再改变前缀，前缀由筛选类型决定
        pass
    
    
    def create_video_tool(self):
        """创建种草助手（视频封面提取）工具页面"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 上方按钮区域
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        import_btn = QPushButton("导入")
        import_btn.clicked.connect(self.import_video_files)
        button_layout.addWidget(import_btn)
        
        import_folder_btn = QPushButton("导入文件夹")
        import_folder_btn.clicked.connect(self.import_video_folder)
        button_layout.addWidget(import_folder_btn)
        
        button_layout.addStretch()
        
        export_table_btn = QPushButton("导出表格")
        export_table_btn.clicked.connect(self.export_video_table)
        button_layout.addWidget(export_table_btn)
        
        export_cover_btn = QPushButton("导出封面")
        export_cover_btn.clicked.connect(self.export_video_covers)
        button_layout.addWidget(export_cover_btn)
        
        layout.addLayout(button_layout)
        
        # 视频文件列表
        self.video_files_table = QTableWidget()
        self.video_files_table.setColumnCount(2)
        self.video_files_table.setHorizontalHeaderLabels(["文件夹名称", "视频标题"])
        self.video_files_table.horizontalHeader().setStretchLastSection(True)
        self.video_files_table.setColumnWidth(0, 150)
        self.video_files_table.verticalHeader().setVisible(False)  # 隐藏行号
        layout.addWidget(self.video_files_table)
        
        # 文件信息标签（类似标签箱唛工具）
        file_info_layout = QHBoxLayout()
        file_info_layout.setContentsMargins(0, 5, 0, 3)
        file_info_layout.setSpacing(0)
        
        self.video_info_label = QLabel("未加载数据")
        self.video_info_label.setFont(QFont("Arial", 9))
        file_info_layout.addWidget(self.video_info_label)
        file_info_layout.addStretch()
        
        layout.addLayout(file_info_layout)
        
        # 存储视频文件列表
        self.video_files_list = []
        
        widget.setLayout(layout)
        return widget
    
    def import_video_files(self):
        """导入视频文件"""
        import os
        from PyQt5.QtWidgets import QFileDialog
        
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择视频文件",
            "",
            "视频文件 (*.mp4 *.avi *.mov *.mkv *.flv *.wmv *.webm *.m4v *.3gp *.ts);;所有文件 (*)"
        )
        
        if not file_paths:
            return
        
        video_extensions = ('.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.webm', '.m4v', '.3gp', '.ts')
        
        for file_path in file_paths:
            if file_path.lower().endswith(video_extensions):
                if file_path not in self.video_files_list:
                    self.video_files_list.append(file_path)
        
        # 刷新表格
        self.refresh_video_files_table()
    
    def import_video_folder(self):
        """导入视频文件夹"""
        import os
        from PyQt5.QtWidgets import QFileDialog
        
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "选择包含视频文件的文件夹"
        )
        
        if not folder_path:
            return
        
        video_extensions = ('.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.webm', '.m4v', '.3gp', '.ts')
        
        # 递归扫描文件夹中的所有视频文件
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith(video_extensions):
                    file_path = os.path.join(root, file)
                    if file_path not in self.video_files_list:
                        self.video_files_list.append(file_path)
        
        # 刷新表格
        self.refresh_video_files_table()
    
    def refresh_video_files_table(self):
        """刷新视频文件表格"""
        self.video_files_table.setRowCount(0)
        
        import os
        for idx, file_path in enumerate(self.video_files_list):
            self.video_files_table.insertRow(idx)
            
            # 文件夹名称（上级文件夹）
            folder_name = os.path.basename(os.path.dirname(file_path))
            folder_item = QTableWidgetItem(folder_name)
            self.video_files_table.setItem(idx, 0, folder_item)
            
            # 视频标题（文件名不含扩展名）
            video_name = os.path.splitext(os.path.basename(file_path))[0]
            title_item = QTableWidgetItem(video_name)
            self.video_files_table.setItem(idx, 1, title_item)
        
        # 更新信息标签
        if self.video_files_list:
            self.video_info_label.setText(f"已加载: {len(self.video_files_list)} 个视频文件")
        else:
            self.video_info_label.setText("未加载数据")
    
    def export_video_table(self):
        """导出视频表格为Excel"""
        if not self.video_files_list:
            QMessageBox.warning(self, "警告", "请先导入视频文件")
            return
        
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "警告", "需要安装openpyxl库。请运行: pip install openpyxl")
            return
        
        # 选择保存位置
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存表格", "", "Excel文件 (*.xlsx);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        try:
            # 创建工作簿
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "视频列表"
            
            # 写入表头
            worksheet['A1'] = "文件夹名称"
            worksheet['B1'] = "视频标题"
            
            # 写入数据
            import os
            for idx, file_path in enumerate(self.video_files_list, start=2):
                folder_name = os.path.basename(os.path.dirname(file_path))
                video_name = os.path.splitext(os.path.basename(file_path))[0]
                
                worksheet[f'A{idx}'] = folder_name
                worksheet[f'B{idx}'] = video_name
            
            # 调整列宽
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 30
            
            # 保存文件
            workbook.save(file_path)
            
            QMessageBox.information(self, "成功", f"表格已导出到: {file_path}")
        
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出表格失败: {str(e)}")
    
    def export_video_covers(self):
        """导出视频封面"""
        if not self.video_files_list:
            QMessageBox.warning(self, "警告", "请先导入视频文件")
            return
        
        # 选择输出文件夹
        output_folder = QFileDialog.getExistingDirectory(
            self, "选择输出文件夹", ""
        )
        
        if not output_folder:
            return
        
        try:
            from main.video.cover_extractor import VideoCoverExtractor
            
            success_count = 0
            failed_count = 0
            error_messages = []
            
            for idx, video_path in enumerate(self.video_files_list):
                try:
                    # 获取视频文件名（不含扩展名）
                    import os
                    video_name = os.path.splitext(os.path.basename(video_path))[0]
                    
                    # 输出图片路径
                    output_image = os.path.join(output_folder, f"{video_name}.jpg")
                    
                    # 提取封面
                    if VideoCoverExtractor.extract_cover(video_path, output_image):
                        success_count += 1
                    else:
                        failed_count += 1
                        error_messages.append(f"无法提取: {video_name}")
                
                except Exception as e:
                    print(f"[ERROR] 处理视频失败: {video_path}, {str(e)}")
                    failed_count += 1
                    error_messages.append(f"错误: {video_name} - {str(e)}")
            
            # 显示结果
            result_msg = f"成功: {success_count} 个\n失败: {failed_count} 个"
            
            if failed_count > 0:
                if failed_count <= 5:
                    result_msg += "\n\n失败原因:\n" + "\n".join(error_messages[:5])
                else:
                    result_msg += "\n\n失败原因 (前5个):\n" + "\n".join(error_messages[:5])
                    result_msg += f"\n... 还有 {failed_count - 5} 个失败"
                
                # 如果全部失败，显示详细错误
                if success_count == 0:
                    result_msg += "\n\n提示: 请确保已安装 ffmpeg\n"
                    result_msg += "macOS: brew install ffmpeg\n"
                    result_msg += "Windows: https://ffmpeg.org/download.html"
            
            QMessageBox.information(self, "导出完成", result_msg)
        
        except ImportError:
            QMessageBox.warning(
                self, 
                "错误", 
                "需要安装 ffmpeg。\n\n"
                "macOS: brew install ffmpeg\n"
                "Windows: https://ffmpeg.org/download.html"
            )
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def create_label_tool(self):
        """创建标签箱唛工具页面"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 第一行：导入按钮和工作表选择
        top_layout = QHBoxLayout()
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(10)
        
        import_excel_btn = QPushButton("导入")
        import_excel_btn.clicked.connect(self.import_excel_file)
        top_layout.addWidget(import_excel_btn)
        
        top_layout.addSpacing(10)
        
        # 工作表选择
        top_layout.addWidget(QLabel("工作表:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.addItem("未加载数据")
        self.sheet_combo.setMinimumWidth(150)
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_changed)
        top_layout.addWidget(self.sheet_combo)
        
        top_layout.addStretch()
        
        add_label_btn = QPushButton("新增标签")
        add_label_btn.clicked.connect(self.show_label_dialog)
        top_layout.addWidget(add_label_btn)
        
        add_box_btn = QPushButton("新增箱唛")
        add_box_btn.clicked.connect(self.show_box_dialog)
        top_layout.addWidget(add_box_btn)
        
        settings_btn = QPushButton("设置")
        settings_btn.clicked.connect(self.show_sheet_mapping_dialog)
        top_layout.addWidget(settings_btn)
        
        export_btn = QPushButton("导出")
        export_btn.clicked.connect(self.export_label_box)
        top_layout.addWidget(export_btn)
        
        layout.addLayout(top_layout)
        
        # 预览下方的选项
        preview_options_layout = QHBoxLayout()
        preview_options_layout.setContentsMargins(0, 0, 0, 0)
        preview_options_layout.setSpacing(10)
        
        # 类型选择
        preview_options_layout.addWidget(QLabel("类型:"))
        self.preview_type_combo = QComboBox()
        self.preview_type_combo.addItems(["3C", "玩具"])
        self.preview_type_combo.setCurrentText("3C")  # 默认显示 3C
        self.preview_type_combo.setMinimumWidth(80)
        self.preview_type_combo.setMaximumWidth(80)
        preview_options_layout.addWidget(self.preview_type_combo)
        
        preview_options_layout.addSpacing(10)  # 四组之间的间距
        
        # 年份
        preview_options_layout.addWidget(QLabel("年份:"))
        self.export_year_input = QLineEdit()
        self.export_year_input.setPlaceholderText("YYYY")
        self.export_year_input.setMaximumWidth(80)
        import datetime
        current_year = datetime.datetime.now().year
        self.export_year_input.setText(str(current_year))
        preview_options_layout.addWidget(self.export_year_input)
        
        preview_options_layout.addSpacing(10)  # 四组之间的间距
        
        # 日期
        preview_options_layout.addWidget(QLabel("日期:"))
        self.export_date_input = QLineEdit()
        self.export_date_input.setPlaceholderText("MMDD")
        self.export_date_input.setMinimumWidth(80)
        self.export_date_input.setMaximumWidth(80)
        # 默认显示今天的日期
        import datetime
        self.export_date_input.setText(datetime.datetime.now().strftime("%m%d"))
        preview_options_layout.addWidget(self.export_date_input)
        
        preview_options_layout.addSpacing(10)  # 四组之间的间距
        
        # 输出内容复选框
        preview_options_layout.addWidget(QLabel("输出内容:"))
        
        self.output_label_checkbox = QCheckBox("标签")
        self.output_label_checkbox.setChecked(True)
        preview_options_layout.addWidget(self.output_label_checkbox)
        
        self.output_box_checkbox = QCheckBox("箱唛")
        self.output_box_checkbox.setChecked(True)
        preview_options_layout.addWidget(self.output_box_checkbox)
        
        self.output_order_checkbox = QCheckBox("预定表")
        self.output_order_checkbox.setChecked(False)
        self.output_order_checkbox.stateChanged.connect(self.on_order_checkbox_changed)
        preview_options_layout.addWidget(self.output_order_checkbox)
        
        preview_options_layout.addStretch()  # 添加拉伸，让识别设置按钮居右
        
        # 识别设置按钮
        recognize_settings_btn = QPushButton("识别设置")
        recognize_settings_btn.clicked.connect(self.show_recognize_settings_dialog)
        preview_options_layout.addWidget(recognize_settings_btn)
        
        layout.addLayout(preview_options_layout)
        
        # 表格预览
        self.excel_preview_table = QTableWidget()
        self.excel_preview_table.setColumnCount(0)
        layout.addWidget(self.excel_preview_table)
        
        # 导入文件名称显示
        file_info_layout = QHBoxLayout()
        file_info_layout.setContentsMargins(0, 5, 0, 3)
        file_info_layout.setSpacing(0)
        
        self.file_name_label = QLabel("未加载数据")
        self.file_name_label.setFont(QFont("Arial", 9))
        file_info_layout.addWidget(self.file_name_label)
        file_info_layout.addStretch()
        
        layout.addLayout(file_info_layout)
        
        widget.setLayout(layout)
        return widget
    
    def show_recognize_settings_dialog(self):
        """显示识别设置对话框"""
        dialog = RecognizeSettingsDialog(self)
        dialog.exec_()
    
    def on_sheet_changed(self, sheet_name):
        """工作表改变时的处理"""
        if not self.current_excel_workbook or not sheet_name:
            return
        
        try:
            worksheet = self.current_excel_workbook[sheet_name]
            
            # 获取最大列数
            max_col = worksheet.max_column
            
            # 设置表格列数
            self.excel_preview_table.setColumnCount(max_col)
            
            # 设置表头（第一行）
            headers = []
            for col in range(1, max_col + 1):
                header_cell = worksheet.cell(1, col)
                headers.append(str(header_cell.value) if header_cell.value else f"列{col}")
            self.excel_preview_table.setHorizontalHeaderLabels(headers)
            
            # 清空表格数据
            self.excel_preview_table.setRowCount(0)
            
            # 读取数据并填充表格（从第 2 行开始，最多显示 100 行）
            for row_idx, row_num in enumerate(range(2, min(worksheet.max_row + 1, 102))):
                self.excel_preview_table.insertRow(row_idx)
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row_num, col_idx)
                    value = cell.value
                    item = QTableWidgetItem(str(value) if value else "")
                    
                    # 设置单元格背景颜色
                    if cell.fill and cell.fill.start_color:
                        try:
                            color_value = cell.fill.start_color.rgb
                            if color_value and color_value != '00000000':
                                # 移除透明度前缀（如果有）
                                if len(color_value) == 8:
                                    color_value = color_value[2:]
                                bg_color = QColor(f"#{color_value}")
                                item.setBackground(bg_color)
                        except:
                            pass
                    
                    # 设置文字颜色
                    if cell.font and cell.font.color:
                        try:
                            color_value = cell.font.color.rgb
                            if color_value and color_value != '00000000':
                                # 移除透明度前缀（如果有）
                                if len(color_value) == 8:
                                    color_value = color_value[2:]
                                text_color = QColor(f"#{color_value}")
                                item.setForeground(text_color)
                        except:
                            pass
                    
                    self.excel_preview_table.setItem(row_idx, col_idx - 1, item)
            
            # 自动适应列宽
            self.excel_preview_table.resizeColumnsToContents()
        except Exception as e:
            print(f"[DEBUG] 工作表切换失败: {str(e)}")
    
    def on_order_checkbox_changed(self):
        """预定表复选框改变时的处理"""
        if self.output_order_checkbox.isChecked():
            # 当预定表被勾选时，清空标签和箱唛的勾选
            self.output_label_checkbox.setChecked(False)
            self.output_box_checkbox.setChecked(False)
    
    def show_label_dialog(self):
        """显示标签生成对话框"""
        dialog = LabelDialog(self)
        dialog.exec_()
    
    def show_box_dialog(self):
        """显示箱唛生成对话框"""
        dialog = BoxDialog(self)
        dialog.exec_()
    
    def download_template(self):
        """下载模板"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(self, "警告", "需要安装openpyxl库。请运行: pip install openpyxl")
            return
        
        # 选择保存位置
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存模板", os.path.expanduser("~/Desktop/标签模板.xlsx"), "Excel文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # 创建新的工作簿
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "标签数据"
            
            # 添加表头
            headers = ["品牌", "产品名称", "SKU", "69码", "SN前缀", "序号", "年份", "日期", "批次", "文件名"]
            worksheet.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 10
            worksheet.column_dimensions['F'].width = 10
            worksheet.column_dimensions['G'].width = 10
            worksheet.column_dimensions['H'].width = 10
            worksheet.column_dimensions['I'].width = 10
            worksheet.column_dimensions['J'].width = 20
            
            # 保存文件
            workbook.save(file_path)
            QMessageBox.information(self, "成功", f"模板已下载到: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"下载模板失败: {str(e)}")
    
    def export_label_box(self):
        """导出标签和箱唛文件"""
        print("[DEBUG] 开始导出...")
        
        # 检查是否导入了文件
        if not self.current_excel_file or not self.current_excel_workbook:
            QMessageBox.warning(self, "警告", "请先导入Excel文件")
            return
        
        # 检查是否选择了类型
        label_type = self.preview_type_combo.currentText()
        if not label_type:
            QMessageBox.warning(self, "警告", "请先选择类型")
            return
        
        print(f"[DEBUG] label_type={label_type}")
        
        # 检查是否输入了日期
        export_date = self.export_date_input.text().strip()
        if not export_date:
            QMessageBox.warning(self, "警告", "请输入日期（MMDD 格式）")
            return
        
        print(f"[DEBUG] export_date={export_date}")
        
        # 检查是否输入了年份
        export_year = self.export_year_input.text().strip()
        if not export_year:
            QMessageBox.warning(self, "警告", "请输入年份（YYYY 格式）")
            return
        
        print(f"[DEBUG] export_year={export_year}")
        
        # 检查是否勾选了输出内容
        export_label = self.output_label_checkbox.isChecked()
        export_box = self.output_box_checkbox.isChecked()
        export_order = self.output_order_checkbox.isChecked()
        
        print(f"[DEBUG] export_label={export_label}, export_box={export_box}, export_order={export_order}")
        
        if not export_label and not export_box and not export_order:
            QMessageBox.warning(self, "警告", "请至少勾选一个输出内容")
            return
        
        try:
            desktop_path = os.path.expanduser("~/Desktop")
            print(f"[DEBUG] desktop_path={desktop_path}")
            
            # 如果同时勾选标签和箱唛，放在同一个文件夹下
            if export_label and export_box:
                parent_folder_name = f"{export_date}-{label_type}标签箱唛"
                parent_folder = os.path.join(desktop_path, parent_folder_name)
                
                if not os.path.exists(parent_folder):
                    os.makedirs(parent_folder)
                
                # 标签文件夹
                label_folder_name = f"{export_date}-{label_type}标签"
                label_folder = os.path.join(parent_folder, label_folder_name)
                
                if not os.path.exists(label_folder):
                    os.makedirs(label_folder)
                
                print(f"[DEBUG] 开始生成标签到: {label_folder}")
                self.generate_labels_from_workbook(label_folder, label_type, export_date, export_year)
                print(f"[DEBUG] 标签生成完成")
                
                # 箱唛文件夹
                box_folder_name = f"{export_date}-{label_type}箱唛"
                box_folder = os.path.join(parent_folder, box_folder_name)
                
                if not os.path.exists(box_folder):
                    os.makedirs(box_folder)
                
                print(f"[DEBUG] 开始生成箱唛到: {box_folder}")
                self.generate_carton_marks_from_workbook(box_folder, label_type, export_date, export_year)
                print(f"[DEBUG] 箱唛生成完成")
            else:
                # 导出标签
                if export_label:
                    folder_name = f"{export_date}-{label_type}标签"
                    export_folder = os.path.join(desktop_path, folder_name)
                    
                    if not os.path.exists(export_folder):
                        os.makedirs(export_folder)
                    
                    print(f"[DEBUG] 开始生成标签到: {export_folder}")
                    self.generate_labels_from_workbook(export_folder, label_type, export_date, export_year)
                    print(f"[DEBUG] 标签生成完成")
                
                # 导出箱唛
                if export_box:
                    folder_name = f"{export_date}-{label_type}箱唛"
                    export_folder = os.path.join(desktop_path, folder_name)
                    
                    if not os.path.exists(export_folder):
                        os.makedirs(export_folder)
                    
                    print(f"[DEBUG] 开始生成箱唛到: {export_folder}")
                    self.generate_carton_marks_from_workbook(export_folder, label_type, export_date, export_year)
                    print(f"[DEBUG] 箱唛生成完成")
            
            # 导出预定表
            if export_order:
                print(f"[DEBUG] 开始生成预定表")
                self.generate_order_form(label_type, export_date, export_year)
                print(f"[DEBUG] 预定表生成完成")
            
            print("[DEBUG] 导出完成")
            QMessageBox.information(self, "成功", "导出完成")
        
        except Exception as e:
            print(f"[DEBUG] 导出出错: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"导出出错: {str(e)}")
    
    def generate_labels_from_workbook(self, export_folder, label_type, export_date, export_year):
        """从工作表中读取数据并生成标签（使用 default 模板）"""
        try:
            print(f"[DEBUG] generate_labels_from_workbook 开始: export_folder={export_folder}, label_type={label_type}")
            
            import shutil
            import pathlib
            import re
            from main.label.label_generator import LabelGenerator
            
            # 保存原始的螺旋桨映射，用于后续比较
            original_propeller_mapping = dict(self.sheet_mapping.get("螺旋桨", {}))
            new_propeller_mappings = {}  # 追踪新添加的映射
            
            # 获取 default 模板路径
            gui_dir = pathlib.Path(__file__).resolve().parent
            default_templates = {
                "label_12": gui_dir / "templates" / "label_12.pld",
                "label_13": gui_dir / "templates" / "label_13.pld",
                "label_14": gui_dir / "templates" / "label_14.pld",
                "label_15": gui_dir / "templates" / "label_15.pld",
                "label_16": gui_dir / "templates" / "label_16.pld",
                "label_propeller": gui_dir / "templates" / "label_propeller.pld",
            }
            
            print(f"[DEBUG] sheet_mapping keys: {list(self.sheet_mapping.keys())}")
            print(f"[DEBUG] label_type in sheet_mapping: {label_type in self.sheet_mapping}")
            
            # 获取对应类型的工作表
            if label_type not in self.sheet_mapping:
                print(f"[DEBUG] label_type 不在 sheet_mapping 中")
                return
            
            brands = self.sheet_mapping[label_type]
            print(f"[DEBUG] brands: {brands}")
            
            total_count = 0
            generated_count = 0
            skipped_rows = []
            
            # 螺旋桨使用 3C 的品牌列表
            if label_type == "螺旋桨":
                brands = self.sheet_mapping.get("3C", {})
                print(f"[DEBUG] 螺旋桨使用 3C 品牌: {brands}")
            
            # 第一步：预扫描，收集所有新的螺旋桨商品编号
            new_propeller_skus = []
            propeller_mapping = self.sheet_mapping.get("螺旋桨", {})
            
            # 总是扫描所有工作表中的螺旋桨商品编号（69码为空且商品编号是纯数字的行）
            print(f"[DEBUG] 开始预扫描螺旋桨商品编号...")
            scan_brands = self.sheet_mapping.get("3C", {}) if label_type == "螺旋桨" else self.sheet_mapping.get(label_type, {})
            
            for brand, sheets in scan_brands.items():
                for sheet_name in sheets:
                    if sheet_name not in self.current_excel_workbook.sheetnames:
                        continue
                    
                    ws = self.current_excel_workbook[sheet_name]
                    for row in range(2, ws.max_row + 1):
                        sku = str(ws.cell(row, 1).value or "").strip()
                        code69 = str(ws.cell(row, 3).value or "").strip()
                        
                        if not sku:
                            continue
                        
                        # 判断是否为螺旋桨：69码为空 且 商品编号是纯数字
                        is_propeller = (not code69 or not code69.startswith("69")) and sku.isdigit()
                        
                        if is_propeller and sku not in propeller_mapping and sku not in new_propeller_skus:
                            new_propeller_skus.append(sku)
                            print(f"[DEBUG] 发现新的螺旋桨商品编号: {sku}")
            
            # 如果有新的螺旋桨商品编号，弹出对话框让用户输入文件名
            if new_propeller_skus:
                print(f"[DEBUG] 有 {len(new_propeller_skus)} 个新的螺旋桨商品编号，弹出对话框")
                # 获取已有的文件名集合（用于检查重名）
                existing_filenames = set(propeller_mapping.values())
                dialog = PropellerFilenameDialog(self, new_propeller_skus, existing_filenames)
                ok = dialog.exec() == QDialog.DialogCode.Accepted
                
                if ok:
                    filenames = dialog.get_filenames()
                    # 保存新的映射
                    if "螺旋桨" not in self.sheet_mapping:
                        self.sheet_mapping["螺旋桨"] = {}
                    
                    for sku, filename in filenames.items():
                        if filename:  # 只保存非空的文件名
                            self.sheet_mapping["螺旋桨"][sku] = filename
                            new_propeller_mappings[sku] = filename
                            print(f"[DEBUG] 保存新的螺旋桨映射: {sku} -> {filename}")
                    
                    self.save_sheet_mapping()
                else:
                    print(f"[DEBUG] 用户取消了文件名输入")
                    return
            
            print(f"[DEBUG] 开始遍历品牌...")
            for brand, sheets in brands.items():
                print(f"[DEBUG] 处理品牌: {brand}, sheets: {sheets}")
                
                for sheet_name in sheets:
                    if sheet_name not in self.current_excel_workbook.sheetnames:
                        print(f"[DEBUG] 工作表不存在: {sheet_name}")
                        continue
                    
                    ws = self.current_excel_workbook[sheet_name]
                    print(f"[DEBUG] 处理工作表: {sheet_name}")
                    
                    if len(sheets) > 1:
                        sheet_folder = os.path.join(export_folder, sheet_name)
                    else:
                        sheet_folder = os.path.join(export_folder, brand)
                    
                    if not os.path.exists(sheet_folder):
                        os.makedirs(sheet_folder)
                    
                    for row in range(2, ws.max_row + 1):
                        # 检查是否需要跳过红色文字行
                        rule_key = f"{brand}|{sheet_name}"
                        should_skip_red = self.sheet_rules.get(label_type, {}).get(rule_key) == "skip_red_text"
                        
                        print(f"[DEBUG] 行{row}：rule_key={rule_key}, should_skip_red={should_skip_red}")
                        
                        if should_skip_red:
                            # 检查该行是否为红色文字
                            sku_cell = ws.cell(row, 1)
                            is_red = False
                            
                            if sku_cell.font and sku_cell.font.color:
                                color = sku_cell.font.color
                                print(f"[DEBUG] 行{row}：color={color}, color.rgb={getattr(color, 'rgb', None)}, color.type={getattr(color, 'type', None)}")
                                
                                # 检查是否为红色
                                if hasattr(color, 'rgb') and color.rgb:
                                    color_str = str(color.rgb).upper()
                                    print(f"[DEBUG] 行{row}：color_str={color_str}")
                                    # 红色的 RGB 值：FF0000 或 FFFF0000
                                    if 'FF0000' in color_str or color_str == 'FFFF0000':
                                        is_red = True
                                elif hasattr(color, 'index') and color.index:
                                    # 检查是否使用了主题颜色或索引颜色
                                    print(f"[DEBUG] 行{row}：color.index={color.index}")
                                    # 红色通常是索引 10 或主题颜色
                                    if color.index == 10 or color.index == 3:
                                        is_red = True
                            
                            if is_red:
                                print(f"[DEBUG] 行{row}：跳过红色文字行")
                                continue
                        
                        # 读取行数据
                        sku = str(ws.cell(row, 1).value or "").strip()  # A列：商品编号（SKU）
                        brand_cell = str(ws.cell(row, 2).value or "").strip()  # B列：品牌
                        code69 = str(ws.cell(row, 3).value or "").strip()  # C列：69码
                        sn_full = str(ws.cell(row, 4).value or "").strip()  # D列：SN码
                        template_filename = str(ws.cell(row, 5).value or "").strip()  # E列：编号
                        
                        if not sku or not template_filename:
                            continue
                        
                        total_count += 1
                        
                        # 判断是否为螺旋桨：69码为空 且 商品编号是纯数字
                        is_propeller = (not code69 or not code69.startswith("69")) and sku.isdigit()
                        
                        try:
                            if is_propeller:
                                # 螺旋桨标签
                                # 编号就是产品名称本身（无序号）
                                product_name = template_filename
                                seq_num = "1"
                                
                                # 使用 label_propeller 模板
                                template_path = default_templates["label_propeller"]
                                if not template_path.exists():
                                    print(f"[DEBUG] 螺旋桨模板不存在: {template_path}")
                                    skipped_rows.append(sku)
                                    continue
                                
                                generator = LabelGenerator(str(template_path))
                                generator.set_label_data(brand_cell, product_name, sku, "", "", export_date)
                                
                                # 根据商品编号和映射获取文件名
                                propeller_mapping = self.sheet_mapping.get("螺旋桨", {})
                                
                                # 此时所有新的螺旋桨商品编号已经在预扫描阶段处理过了
                                if sku not in propeller_mapping:
                                    print(f"[DEBUG] 行{row}：螺旋桨商品编号 {sku} 未在映射中，跳过")
                                    skipped_rows.append(sku)
                                    continue
                                
                                filename = propeller_mapping[sku]
                                
                                output_filename = f"{filename}.pld"
                                output_path = os.path.join(sheet_folder, output_filename)
                                
                                # 生成文件
                                if generator.generate_pld(sheet_folder, seq_num=seq_num):
                                    # 重命名为最终文件名
                                    default_generated = os.path.join(sheet_folder, f"{seq_num}.{product_name}.pld")
                                    if os.path.exists(default_generated):
                                        shutil.move(default_generated, output_path)
                                    
                                    generated_count += 1
                                    print(f"[DEBUG] 生成螺旋桨标签: {output_path}")
                                else:
                                    print(f"[DEBUG] 行{row}：生成螺旋桨标签失败")
                                    skipped_rows.append(sku)
                            else:
                                # 3C/玩具标签
                                # 编号格式：序号.产品名称（如 "180.JD-H36灰色+VR+1"）或 序号产品名称（如 "60JDx-901MaX+1"）
                                import re
                                
                                if "." in template_filename:
                                    # 有点号的格式：序号.产品名称
                                    parts = template_filename.split(".", 1)
                                    seq_num = parts[0]
                                    product_name = parts[1]
                                else:
                                    # 没有点号的格式：尝试用正则提取序号
                                    # 序号是开头的数字，后面跟着非数字字符
                                    match = re.match(r'^(\d+)([A-Za-z].*)', template_filename)
                                    if match:
                                        seq_num = match.group(1)
                                        product_name = match.group(2)
                                    else:
                                        # 无法识别序号，使用默认值
                                        seq_num = "1"
                                        product_name = template_filename
                                
                                print(f"[DEBUG] 行{row}：template_filename={template_filename}, seq_num={seq_num}, product_name={product_name}")
                                
                                # 从 D列 的 SN 码中提取信息
                                # SN 格式: SG/NB + 序号(1-3位) + 年份(2或4位) + 日期(4位) + 批次(3位)
                                if not sn_full:
                                    print(f"[DEBUG] 行{row}：D列SN码为空")
                                    skipped_rows.append(sku)
                                    continue
                                
                                sn_parts = self.parse_sn(sn_full)
                                
                                if not sn_parts:
                                    print(f"[DEBUG] 行{row}：无法解析SN码: {sn_full}")
                                    skipped_rows.append(sku)
                                    continue
                                
                                sn_prefix = sn_parts['prefix']
                                sn_date = sn_parts['date']
                                
                                # 选择合适的模板（根据 SN 前缀）
                                if sn_prefix == "NB":
                                    template_key = "label_13"  # 玩具用 label_13
                                else:
                                    template_key = "label_12"  # 3C 用 label_12
                                
                                template_path = default_templates[template_key]
                                if not template_path.exists():
                                    print(f"[DEBUG] 模板不存在: {template_path}")
                                    skipped_rows.append(sku)
                                    continue
                                
                                generator = LabelGenerator(str(template_path))
                                generator.set_label_data(brand_cell, product_name, sku, code69, sn_full, sn_date)
                                
                                output_filename = f"{template_filename}.pld"
                                output_path = os.path.join(sheet_folder, output_filename)
                                
                                # 生成文件（不使用序号前缀）
                                if generator.generate_pld(sheet_folder, seq_num=seq_num, use_seq_prefix=False):
                                    # 获取生成的文件名（product_name.pld）
                                    generated_filename = f"{product_name}.pld"
                                    generated_path = os.path.join(sheet_folder, generated_filename)
                                    
                                    # 如果生成的文件名与目标文件名不同，则重命名
                                    if generated_path != output_path and os.path.exists(generated_path):
                                        shutil.move(generated_path, output_path)
                                    
                                    generated_count += 1
                                    print(f"[DEBUG] 生成标签: {output_path}")
                                else:
                                    print(f"[DEBUG] 行{row}：生成标签失败")
                                    skipped_rows.append(sku)
                        
                        except Exception as e:
                            print(f"[DEBUG] 行{row}生成标签失败: {str(e)}")
                            skipped_rows.append(sku)
            
            # 显示统计信息
            self.show_export_summary(total_count, generated_count, {}, skipped_rows)
            
            # 检查是否有新的螺旋桨映射被保存
            if new_propeller_mappings:
                msg = "已保存以下新的螺旋桨映射:\n\n"
                for sku, filename in list(new_propeller_mappings.items())[:10]:
                    msg += f"  • {sku} -> {filename}\n"
                if len(new_propeller_mappings) > 10:
                    msg += f"  ... 还有 {len(new_propeller_mappings) - 10} 个\n"
                QMessageBox.information(self, "螺旋桨映射", msg)
        
        except Exception as e:
            print(f"[DEBUG] 生成标签失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def generate_carton_marks_from_workbook(self, export_folder, label_type, export_date, export_year):
        """从工作表中读取数据并生成箱唛"""
        try:
            import re
            
            # 获取箱唛识别工作表名称
            box_recognize_config = self.config_manager.get('box_recognize_config', {
                'sheet_name': '箱唛'
            })
            
            box_sheet_name = box_recognize_config['sheet_name']
            
            # 检查工作表是否存在
            if box_sheet_name not in self.current_excel_workbook.sheetnames:
                found_sheet = None
                for name in self.current_excel_workbook.sheetnames:
                    if name.strip() == box_sheet_name.strip():
                        found_sheet = name
                        break
                
                if not found_sheet:
                    print(f"[DEBUG] 箱唛工作表不存在: {box_sheet_name}")
                    return
                
                box_sheet_name = found_sheet
            
            ws = self.current_excel_workbook[box_sheet_name]
            
            print(f"[DEBUG] 处理箱唛工作表: {box_sheet_name}")
            
            # 第一阶段：扫描所有行，检测重复文件
            print(f"[DEBUG] 第一阶段：扫描重复文件...")
            duplicate_files = set()
            
            for row in range(1, ws.max_row + 1):
                seq_cell = ws.cell(row, 1)
                if not seq_cell.value:
                    continue
                
                try:
                    seq_num = int(str(seq_cell.value).strip())
                except (ValueError, TypeError):
                    continue
                
                b_col = 2
                c_col = 3
                
                required_cells = [
                    (row, c_col),
                    (row + 1, c_col),
                    (row + 2, c_col),
                    (row + 3, c_col)
                ]
                
                skip = False
                values = {}
                
                for check_row, check_col in required_cells:
                    cell = ws.cell(check_row, check_col)
                    if not cell.value:
                        skip = True
                        break
                    
                    cell_value = str(cell.value).strip()
                    if cell_value.startswith("#"):
                        skip = True
                        break
                    
                    values[check_row] = cell_value
                
                if skip:
                    continue
                
                po_number = values[row + 2]
                if not re.match(r'^\d{10}$', po_number):
                    continue
                
                city = values[row]
                city_match = re.match(r'^(.+?)(\d+)?$', city)
                city_base = city_match.group(1) if city_match else city
                city_suffix = city_match.group(2) if city_match else None
                
                city_code = None
                for code, info in CartonMarkGenerator.CITY_OPTIONS.items():
                    if info["label"] == city_base:
                        city_code = code
                        break
                
                if not city_code:
                    continue
                
                if city_suffix:
                    output_filename = f"{city_code}.{city_base}箱唛-{city_suffix}.pld"
                else:
                    output_filename = f"{city_code}.{city_base}箱唛.pld"
                output_path = os.path.join(export_folder, output_filename)
                
                if os.path.exists(output_path):
                    duplicate_files.add(output_path)
            
            # 如果有重复文件，询问用户是否覆盖
            if duplicate_files:
                msg = f"检测到 {len(duplicate_files)} 个重复文件:\n\n"
                for i, file_path in enumerate(sorted(duplicate_files)):
                    if i < 5:
                        msg += f"  • {os.path.basename(file_path)}\n"
                if len(duplicate_files) > 5:
                    msg += f"  ... 还有 {len(duplicate_files) - 5} 个\n"
                msg += f"\n是否全部覆盖?"
                
                reply = QMessageBox.question(self, "重复文件提示", msg, QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.No:
                    QMessageBox.information(self, "已取消", "导出已取消")
                    return
            
            # 第二阶段：生成文件
            print(f"[DEBUG] 第二阶段：生成文件...")
            total_count = 0
            generated_count = 0
            skipped_rows = []
            
            for row in range(1, ws.max_row + 1):
                seq_cell = ws.cell(row, 1)
                if not seq_cell.value:
                    continue
                
                try:
                    seq_num = int(str(seq_cell.value).strip())
                except (ValueError, TypeError):
                    continue
                
                print(f"[DEBUG] 找到序号: {seq_num} (行 {row})")
                total_count += 1
                
                b_col = 2
                c_col = 3
                
                required_cells = [
                    (row, c_col),
                    (row + 1, c_col),
                    (row + 2, c_col),
                    (row + 3, c_col)
                ]
                
                skip = False
                values = {}
                
                for check_row, check_col in required_cells:
                    cell = ws.cell(check_row, check_col)
                    if not cell.value:
                        print(f"[DEBUG] 行{row}: 单元格({check_row},{check_col})为空，跳过")
                        skip = True
                        break
                    
                    cell_value = str(cell.value).strip()
                    
                    if cell_value.startswith("#"):
                        print(f"[DEBUG] 行{row}: 单元格({check_row},{check_col})是公式错误，跳过")
                        skip = True
                        break
                    
                    values[check_row] = cell_value
                
                if skip:
                    skipped_rows.append(f"序号{seq_num}")
                    continue
                
                po_number = values[row + 2]
                if not re.match(r'^\d{10}$', po_number):
                    print(f"[DEBUG] 行{row}: 采购单号'{po_number}'不是10位数字，跳过")
                    skipped_rows.append(f"序号{seq_num}(采购单号格式错误)")
                    continue
                
                city = values[row]
                supplier_code = values[row + 1]
                warehouse = values[row + 3]
                
                print(f"[DEBUG] 行{row}: 城市={city}, 供应商={supplier_code}, 采购单号={po_number}, 库房={warehouse}")
                
                city_match = re.match(r'^(.+?)(\d+)?$', city)
                city_base = city_match.group(1) if city_match else city
                city_suffix = city_match.group(2) if city_match else None
                
                city_code = None
                for code, info in CartonMarkGenerator.CITY_OPTIONS.items():
                    if info["label"] == city_base:
                        city_code = code
                        break
                
                if not city_code:
                    print(f"[DEBUG] 未找到城市'{city_base}'的代码")
                    skipped_rows.append(f"序号{seq_num}(未找到{city}城市代码)")
                    continue
                
                if city_suffix:
                    output_filename = f"{city_code}.{city_base}箱唛-{city_suffix}.pld"
                else:
                    output_filename = f"{city_code}.{city_base}箱唛.pld"
                output_path = os.path.join(export_folder, output_filename)
                
                generator = CartonMarkGenerator()
                
                if generator.generate_to_path(po_number, warehouse, city_code, supplier_code, output_path):
                    generated_count += 1
                    print(f"[DEBUG] 生成箱唛: {output_path}")
                else:
                    print(f"[DEBUG] 生成箱唛失败: {output_path}")
                    skipped_rows.append(f"序号{seq_num}(生成失败)")
            
            # 显示统计信息
            msg = f"箱唛导出统计:\n\n"
            msg += f"应该生成: {total_count} 个\n"
            msg += f"实际生成: {generated_count} 个\n"
            msg += f"被跳过: {len(skipped_rows)} 个\n"
            
            if skipped_rows:
                msg += f"\n被跳过的箱唛:\n"
                for item in skipped_rows[:10]:
                    msg += f"  • {item}\n"
                if len(skipped_rows) > 10:
                    msg += f"  ... 还有 {len(skipped_rows) - 10} 个\n"
            
            QMessageBox.information(self, "箱唛导出完成", msg)
        
        except Exception as e:
            print(f"[DEBUG] 生成箱唛失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def generate_propeller_labels(self, export_folder, export_date):
        """生成螺旋桨标签"""
        try:
            import shutil
            
            # 获取螺旋桨映射
            propeller_mapping = self.sheet_mapping.get("螺旋桨", {})
            
            # 获取模板文件夹
            template_dir = os.path.join(os.path.dirname(__file__), "templates", "标签模板", "螺旋桨")
            
            if not os.path.exists(template_dir):
                print(f"[DEBUG] 螺旋桨模板文件夹不存在: {template_dir}")
                return
            
            # 遍历映射，复制文件
            for sku, template_filename in propeller_mapping.items():
                # 查找模板文件
                template_file = os.path.join(template_dir, template_filename)
                
                if not os.path.exists(template_file):
                    print(f"[DEBUG] 螺旋桨模板文件不存在: {template_file}")
                    continue
                
                # 生成输出文件名
                output_filename = template_filename
                output_path = os.path.join(export_folder, output_filename)
                
                # 检查文件是否已存在
                if os.path.exists(output_path):
                    print(f"[DEBUG] 文件已存在: {output_path}")
                    continue
                
                # 复制模板
                shutil.copy(template_file, output_path)
                
                # 修改模板中的日期
                self.modify_propeller_template(output_path, export_date)
                
                print(f"[DEBUG] 生成螺旋桨标签: {output_path}")
        
        except Exception as e:
            print(f"[DEBUG] 生成螺旋桨标签失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def find_template_file(self, template_dir, product_name):
        """查找对应的模板文件"""
        try:
            if not os.path.exists(template_dir):
                return None
            
            # 遍历模板文件夹，查找包含产品名称的文件
            for filename in os.listdir(template_dir):
                if product_name in filename and filename.endswith('.pld'):
                    return os.path.join(template_dir, filename)
            
            return None
        except Exception as e:
            print(f"[DEBUG] 查找模板文件失败: {str(e)}")
            return None
    
    def find_templates_by_prefix(self, template_dir, prefix):
        """根据前缀查找模板文件"""
        try:
            if not os.path.exists(template_dir):
                return []
            
            matching_files = []
            # 遍历模板文件夹，查找以前缀开头的文件
            for filename in os.listdir(template_dir):
                if filename.startswith(prefix) and filename.endswith('.pld'):
                    # 提取文件名（不含扩展名）
                    name_without_ext = filename[:-4]  # 去掉 .pld
                    matching_files.append(name_without_ext)
            
            return sorted(matching_files)
        except Exception as e:
            print(f"[DEBUG] 查找前缀模板失败: {str(e)}")
            return []
    
    def show_export_summary(self, total_count, generated_count, missing_templates, skipped_rows):
        """显示导出统计信息"""
        # 计算未生成的数量
        not_generated_count = total_count - generated_count - len(skipped_rows) - len(missing_templates)
        
        # 构建消息
        msg = f"导出统计:\n\n"
        msg += f"应该生成: {total_count} 个\n"
        msg += f"实际生成: {generated_count} 个\n"
        msg += f"被跳过: {len(skipped_rows)} 个\n"
        msg += f"未找到模板: {len(missing_templates)} 个\n"
        
        if skipped_rows:
            msg += f"\n被跳过的标签:\n"
            for item in skipped_rows[:10]:  # 最多显示10个
                msg += f"  • {item}\n"
            if len(skipped_rows) > 10:
                msg += f"  ... 还有 {len(skipped_rows) - 10} 个\n"
        
        if missing_templates:
            msg += f"\n未找到模板的标签:\n"
            for excel_name, templates in list(missing_templates.items())[:10]:  # 最多显示10个
                msg += f"  • {excel_name}\n"
            if len(missing_templates) > 10:
                msg += f"  ... 还有 {len(missing_templates) - 10} 个\n"
        
        QMessageBox.information(self, "导出完成", msg)
    
    def modify_template_date(self, template_path, export_date, export_year, label_type=None):
        """修改模板中的日期和SN码中的年份日期"""
        try:
            from main.label.pld_modifier import PLDModifier
            
            modifier = PLDModifier(template_path)
            
            if label_type == "3C":
                # 3C: 年份是2位数字 + 日期4位数字 = 6位总长
                year_2digit = export_year[-2:]  # 取最后2位
                new_datetime = year_2digit + export_date
                
                # 自动检测并替换年份+日期占位符
                modifier.replace_datetime_placeholder(new_datetime, encoding='gbk')
                
                # 替换单独的日期占位符（右上角）
                modifier.replace_date_placeholder(export_date, encoding='gbk')
                    
            elif label_type == "玩具":
                # 玩具: 年份是4位数字 + 日期4位数字 = 8位总长
                new_datetime = export_year + export_date
                
                # 自动检测并替换年份+日期占位符
                modifier.replace_datetime_placeholder(new_datetime, encoding='gbk')
                
                # 替换单独的日期占位符（右上角）
                modifier.replace_date_placeholder(export_date, encoding='gbk')
            
            # 保存修改
            modifier.save(template_path)
        except Exception as e:
            print(f"[DEBUG] 修改模板日期失败: {str(e)}")
    
    def modify_template(self, template_path, sn, export_date, export_year):
        """修改模板中的日期和 SN 码中的日期部分"""
        try:
            from main.label.pld_modifier import PLDModifier
            
            modifier = PLDModifier(template_path)
            
            # 替换 SN 码中的日期部分
            # SN 码的日期是倒数第 7~4 位
            sn_str = str(sn).strip()
            if len(sn_str) >= 7:
                # 构建新的 SN 码，替换日期部分
                new_sn = sn_str[:-7] + export_date + sn_str[-3:]
                modifier.replace_field(sn_str, new_sn, len(sn_str), encoding='gbk')
            
            # 替换模板中的日期（右上角的日期）
            modifier.replace_field("0000", export_date, 4, encoding='gbk')
            
            # 保存修改
            modifier.save(template_path)
        except Exception as e:
            print(f"[DEBUG] 修改模板失败: {str(e)}")
    
    def modify_propeller_template(self, template_path, export_date):
        """修改螺旋桨模板中的日期"""
        try:
            from main.label.pld_modifier import PLDModifier
            
            modifier = PLDModifier(template_path)
            
            # 替换模板中的日期（右上角的日期）
            modifier.replace_field("0000", export_date, 4, encoding='gbk')
            
            # 保存修改
            modifier.save(template_path)
        except Exception as e:
            print(f"[DEBUG] 修改螺旋桨模板失败: {str(e)}")
    
    def is_city_column(self, cell_str):
        """检查单元格是否是城市列（城市名或城市+数字）"""
        if not cell_str:
            return False
        
        # 检查是否包含城市基础名称
        for code, info in CartonMarkGenerator.CITY_OPTIONS.items():
            city_base = info["label"]
            if city_base in cell_str:
                # 检查是否是"城市"或"城市+数字"的格式
                # 例如："北京"、"北京2"
                remainder = cell_str.replace(city_base, "", 1).strip()
                if not remainder or remainder.isdigit():
                    return True
        
        return False
    
    def generate_order_form(self, label_type, export_date, export_year):
        """生成预定表"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime, timedelta
            import re
            
            # 获取对应类型的工作表
            if label_type not in self.sheet_mapping:
                print(f"[DEBUG] label_type 不在 sheet_mapping 中")
                return
            
            brands = self.sheet_mapping[label_type]
            
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "预定表"
            
            # 设置表头
            headers = [
                "商品编号", "销售场ID", "采购渠道ID", "ouid-供应商简码", 
                "开启有限/无限预订", "新增/删除", "配送中心名称", "配送中心编号", 
                "仓编号", "有限预订数量", "有限预订预计到货日期xxxx-xx-xx", 
                "预定超限原因", "备注", "填写时间"
            ]
            
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 收集所有城市及其预计到货日期
            city_delivery_days = {}  # {city_full: days}
            
            # 第一步：扫描所有配货表，收集城市列（连续的城市列）
            all_cities = set()
            city_columns = {}  # {sheet_name: [(city_full, col_index), ...]}
            
            for brand, sheets in brands.items():
                for sheet_name in sheets:
                    if sheet_name not in self.current_excel_workbook.sheetnames:
                        continue
                    
                    ws_source = self.current_excel_workbook[sheet_name]
                    
                    # 从第1行查找城市列（从编号列往右查找）
                    # 假设编号在E列
                    col_idx = 5  # E列
                    cities_in_sheet = []
                    found_first_city = False
                    
                    for col in range(col_idx + 1, ws_source.max_column + 1):
                        cell_value = ws_source.cell(1, col).value
                        if not cell_value:
                            cell_str = ""
                        else:
                            cell_str = str(cell_value).strip()
                        
                        # 检查是否是城市列
                        if self.is_city_column(cell_str):
                            found_first_city = True
                            cities_in_sheet.append((cell_str, col))
                            all_cities.add(cell_str)
                        elif found_first_city:
                            # 已经找到第一个城市，现在遇到非城市，停止读取
                            break
                    
                    if cities_in_sheet:
                        city_columns[sheet_name] = cities_in_sheet
                        print(f"[DEBUG] 工作表 {sheet_name} 的城市列: {[c[0] for c in cities_in_sheet]}")
            
            # 第二步：弹窗让用户填写每个城市的预计到货日期
            if all_cities:
                dialog = OrderFormDeliveryDialog(sorted(all_cities), self)
                if dialog.exec_() == QDialog.Accepted:
                    city_delivery_days = dialog.get_delivery_days()
                else:
                    QMessageBox.information(self, "已取消", "预定表导出已取消")
                    return
            
            # 第三步：生成预定表数据
            current_time = datetime.now()
            
            # 先收集所有数据到列表
            all_rows = []
            
            for brand, sheets in brands.items():
                for sheet_name in sheets:
                    if sheet_name not in self.current_excel_workbook.sheetnames:
                        continue
                    
                    ws_source = self.current_excel_workbook[sheet_name]
                    
                    # 从第2行开始读取数据
                    for row in range(2, ws_source.max_row + 1):
                        # 获取商品编号（A列）
                        product_code = ws_source.cell(row, 1).value
                        if not product_code:
                            continue
                        
                        product_code = str(product_code).strip()
                        
                        # 验证商品编号是否为12位数字
                        if not (len(product_code) == 12 and product_code.isdigit()):
                            continue
                        
                        # 遍历该行的城市列，生成预定表行
                        if sheet_name in city_columns:
                            for city_full, col_idx in city_columns[sheet_name]:
                                quantity = ws_source.cell(row, col_idx).value
                                
                                # 如果数量为空或0，跳过
                                if not quantity:
                                    continue
                                
                                try:
                                    quantity = int(quantity)
                                    if quantity <= 0:
                                        continue
                                except (ValueError, TypeError):
                                    continue
                                
                                # 计算预计到货日期（YYYY/M/D格式）
                                delivery_days = city_delivery_days.get(city_full, 0)
                                delivery_date_obj = current_time + timedelta(days=delivery_days)
                                delivery_date = f"{delivery_date_obj.year}/{delivery_date_obj.month}/{delivery_date_obj.day}"
                                
                                # 填写时间（YYYY/M/D格式）
                                fill_time_formatted = f"{current_time.year}/{current_time.month}/{current_time.day}"
                                
                                # 添加行到列表
                                row_data = [
                                    product_code,           # 商品编号
                                    "",                     # 销售场ID
                                    "1001",                 # 采购渠道ID（固定1001）
                                    "",                     # ouid-供应商简码（不填）
                                    "有限预订",             # 开启有限/无限预订
                                    "新增",                 # 新增/删除
                                    city_full,              # 配送中心名称（城市名）
                                    "",                     # 配送中心编号（不填）
                                    "",                     # 仓编号
                                    quantity,               # 有限预订数量
                                    delivery_date,          # 有限预订预计到货日期
                                    "",                     # 预定超限原因
                                    "",                     # 备注
                                    fill_time_formatted     # 填写时间
                                ]
                                
                                all_rows.append(row_data)
            
            # 按城市排序（第7列是配送中心名称）
            all_rows.sort(key=lambda x: x[6])
            
            # 将排序后的数据写入工作表
            for row_data in all_rows:
                ws.append(row_data)
            
            # 调整列宽 - 只有商品编号列(A)适应文字，其他列保持默认
            from openpyxl.utils import get_column_letter
            
            # 只调整A列（商品编号）的列宽
            col_letter = 'A'
            max_length = 0
            
            for cell in ws[col_letter]:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # 根据内容自动调整A列列宽，加2作为边距
            adjusted_width = max_length + 2
            # 最小宽度为10
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[col_letter].width = adjusted_width
            
            # 保存文件到桌面
            desktop_path = os.path.expanduser("~/Desktop")
            output_filename = f"{export_date}-{label_type}预定表.xlsx"
            output_path = os.path.join(desktop_path, output_filename)
            
            wb.save(output_path)
            print(f"[DEBUG] 预定表已生成: {output_path}")
            QMessageBox.information(self, "成功", f"预定表已导出到: {output_filename}")
        
        except Exception as e:
            print(f"[DEBUG] 生成预定表失败: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"生成预定表失败: {str(e)}")
    
    
    def load_sheet_mapping(self):
        """从配置文件加载映射关系"""
        import json
        default_mapping = self.init_sheet_mapping()
        
        try:
            saved_mapping = self.config_manager.get('sheet_mapping')
            if saved_mapping:
                return saved_mapping
        except Exception as e:
            print(f"[DEBUG] 加载映射失败: {str(e)}")
        
        return default_mapping
    
    def save_sheet_mapping(self):
        """保存映射关系到配置文件"""
        try:
            self.config_manager.set('sheet_mapping', self.sheet_mapping)
            self.config_manager.save_config()
            return True
        except Exception as e:
            print(f"[DEBUG] 保存映射失败: {str(e)}")
            return False
    
    def load_sheet_rules(self):
        """从配置文件加载规则关系"""
        try:
            saved_rules = self.config_manager.get('sheet_rules')
            if saved_rules:
                return saved_rules
        except Exception as e:
            print(f"[DEBUG] 加载规则失败: {str(e)}")
        
        # 返回默认规则
        return self.init_sheet_rules()
    
    def init_sheet_rules(self):
        """初始化默认规则关系"""
        rules = {}
        
        # 遍历映射关系，为包含"拆2"的工作表设置默认规则
        for label_type, brands in self.sheet_mapping.items():
            if label_type == "螺旋桨":
                continue
            
            rules[label_type] = {}
            for brand, sheets in brands.items():
                if isinstance(sheets, list):
                    for sheet in sheets:
                        # 如果工作表名称包含"拆2"，设置默认规则为"跳过红色文字行"
                        if "拆2" in sheet:
                            rule_key = f"{brand}|{sheet}"
                            rules[label_type][rule_key] = "skip_red_text"
        
        return rules
    
    def save_sheet_rules(self):
        """保存规则关系到配置文件"""
        try:
            self.config_manager.set('sheet_rules', self.sheet_rules)
            self.config_manager.save_config()
            return True
        except Exception as e:
            print(f"[DEBUG] 保存规则失败: {str(e)}")
            return False
    
    def init_sheet_mapping(self):
        """初始化品牌-工作表映射关系"""
        return {
            "3C": {
                "外星人": ["外星人配货表"],
                "三只梨": ["梨配货表"],
                "兽": ["兽配货表"],
                "兽无人机": ["兽无人机拆1", "兽无人机拆2"]
            },
            "玩具": {
                "外星人": ["外仓库配货表"],
                "兽模型": ["兽仓库配货表"]
            },
            "螺旋桨": {}  # 螺旋桨映射（商品编号 -> 模板文件名）
        }
    
    def show_sheet_mapping_dialog(self):
        """显示工作表映射设置对话框"""
        dialog = SheetMappingDialog(self, self.sheet_mapping)
        if dialog.exec_() == QDialog.Accepted:
            self.sheet_mapping = dialog.get_mapping()
            self.sheet_rules = dialog.get_rules()
            self.save_sheet_mapping()
            self.save_sheet_rules()
    
    
    def check_sheet_mapping(self, label_type, workbook):
        """检查导入的Excel文件中是否包含所有设置的工作表"""
        if not label_type or label_type not in self.sheet_mapping:
            return
        
        # 获取所有非隐藏的工作表名称
        available_sheets = [sheet for sheet in workbook.sheetnames if not workbook[sheet].sheet_state == 'hidden']
        
        # 获取该类型下所有需要的工作表
        required_sheets = []
        for brand, sheets in self.sheet_mapping[label_type].items():
            required_sheets.extend(sheets)
        
        # 检查是否找到所有需要的工作表
        missing_sheets = [sheet for sheet in required_sheets if sheet not in available_sheets]
        found_sheets = [sheet for sheet in required_sheets if sheet in available_sheets]
        
        if missing_sheets:
            msg = f"类型: {label_type}\n\n找到的工作表:\n" + "\n".join(found_sheets) if found_sheets else "未找到任何工作表"
            msg += f"\n\n缺少的工作表:\n" + "\n".join(missing_sheets)
            QMessageBox.warning(self, "工作表检查", msg)
        else:
            QMessageBox.information(self, "工作表检查", f"类型: {label_type}\n\n已识别到所有设置的工作表")
    
    
    
    def generate_box(self):
        """创建工具栏"""
        toolbar_layout = QHBoxLayout()
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        toolbar_layout.setSpacing(10)
        
        # 加载文件按钮
        load_btn = QPushButton("导入")
        load_btn.clicked.connect(self.load_file)
        toolbar_layout.addWidget(load_btn)
        
        toolbar_layout.addSpacing(10)
        
        # 推广计划选择
        toolbar_layout.addWidget(QLabel("推广计划:"))
        self.plan_combo = QComboBox()
        self.plan_combo.addItem("未加载数据")  # 初始化时添加默认项
        self.plan_combo.setMinimumWidth(150)
        self.plan_combo.currentTextChanged.connect(self.on_plan_changed)
        toolbar_layout.addWidget(self.plan_combo)
        
        toolbar_layout.addStretch()
        
        # 设置按钮
        settings_btn = QPushButton("设置")
        settings_btn.clicked.connect(self.open_settings)
        toolbar_layout.addWidget(settings_btn)
        
        # 导出按钮
        export_btn = QPushButton("导出")
        export_btn.clicked.connect(self.export_data)
        toolbar_layout.addWidget(export_btn)
        
        return toolbar_layout
    
    def create_keywords_table(self):
        """创建关键词分析表"""
        self.keywords_table = QTableWidget()
        self.keywords_table.setColumnCount(8)
        self.keywords_table.setHorizontalHeaderLabels([
            "关键词", "展现数", "花费", "点击率(%)", "转化率(%)", "订单数", "订单金额(¥)", "优化策略"
        ])
        
        # 设置表格为只读
        self.keywords_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # 显示网格线
        self.keywords_table.setShowGrid(True)
        
        # 隐藏序号列
        self.keywords_table.verticalHeader().setVisible(False)
        
        # 设置列宽
        header = self.keywords_table.horizontalHeader()
        # 第1列（关键词）设置宽度
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        self.keywords_table.setColumnWidth(0, 120)
        
        # 第2到第7列使用统一的固定宽度
        for col in range(1, 7):
            header.setSectionResizeMode(col, QHeaderView.Fixed)
            self.keywords_table.setColumnWidth(col, 75)
        
        # 第8列（优化策略）自动拉伸
        header.setSectionResizeMode(7, QHeaderView.Stretch)
    
    def load_file(self):
        """加载 CSV 文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择关键词报表", self.last_file_path, "CSV 文件 (*.csv);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        # 保存文件所在的目录路径
        self.last_file_path = os.path.dirname(file_path)
        self.config_manager.set('last_csv_path', self.last_file_path)
        self.config_manager.save_config()
        
        data, errors = self.processor.load_keyword_report(file_path)
        
        if errors:
            QMessageBox.warning(self, "警告", f"加载完成，但有 {len(errors)} 个警告")
        
        if data:
            self.optimizer.load_data(data)
            self.current_keywords = self.optimizer.analyze_all_keywords()
            
            # 计算推荐阈值（保存到临时变量，在阈值设置中显示）
            self.recommended_thresholds = self.optimizer.get_recommended_thresholds()
            
            # 更新推广计划下拉框
            self.update_plan_combo()
            
            # 显示第一个推广计划的数据
            if self.plan_combo.count() > 0:
                self.plan_combo.setCurrentIndex(0)
                # 手动触发 on_plan_changed
                self.on_plan_changed(self.plan_combo.currentText())
            
            # 更新info_label显示文件名
            file_name = os.path.basename(file_path)
            self.info_label.setText(f"已加载: {file_name}")
        else:
            QMessageBox.critical(self, "加载失败", "没有有效数据")
    
    def update_plan_combo(self):
        """更新推广计划下拉框"""
        self.plan_combo.blockSignals(True)
        self.plan_combo.clear()
        
        # 获取所有推广计划
        plans = set()
        for kw in self.current_keywords:
            plan = kw.get('推广计划', '')
            if plan:  # 只添加非空的推广计划
                plans.add(plan)
        
        # 如果没有找到推广计划，使用默认值
        if not plans:
            plans = {'Uncategorized'}
        
        self.plan_combo.addItems(sorted(plans))
        self.plan_combo.blockSignals(False)
        
        # 手动触发信号以显示第一个推广计划
        if self.plan_combo.count() > 0:
            self.plan_combo.setCurrentIndex(0)
    
    def on_plan_changed(self, plan_name):
        """推广计划改变时的处理"""
        if not plan_name or not self.current_keywords:
            return
        
        # 筛选该推广计划的关键词
        filtered_keywords = [kw for kw in self.current_keywords if kw.get('推广计划', '') == plan_name]
        
        # 如果没有找到匹配的关键词，显示所有关键词
        if not filtered_keywords:
            filtered_keywords = self.current_keywords
        
        # 更新表格
        self.update_keywords_table(filtered_keywords)
        
        # 更新统计信息
        total_cost = sum(kw.get('花费', 0) for kw in filtered_keywords)
        total_revenue = sum(kw.get('总订单金额', 0) for kw in filtered_keywords)
        roi = (total_revenue / total_cost) if total_cost > 0 else 0
        
        self.info_label.setText(
            f"关键词数: {len(filtered_keywords)}          "
            f"总花费: ¥{total_cost:.2f}          "
            f"总订单金额: ¥{total_revenue:.2f}          "
            f"投产比: {roi:.2f}"
        )
    
    def update_keywords_table(self, keywords=None):
        """更新关键词表"""
        if keywords is None:
            keywords = self.current_keywords
        
        self.keywords_table.setRowCount(len(keywords))
        
        for row, keyword in enumerate(keywords):
            items = [
                keyword.get('关键词', ''),
                str(keyword.get('展现数', 0)),
                f"{keyword.get('花费', 0):.2f}",
                f"{keyword.get('点击率(%)', 0):.2f}",
                f"{keyword.get('转化率(%)', 0):.2f}",
                str(keyword.get('总订单行', 0)),
                f"{keyword.get('总订单金额', 0):.2f}",
                keyword.get('优化策略', '')
            ]
            
            for col, item in enumerate(items):
                table_item = QTableWidgetItem(item)
                
                # 为所有单元格添加 tooltip
                table_item.setToolTip(item)
                
                # 第2到第7列（数值列）居右对齐
                if col >= 1 and col <= 6:
                    table_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                # 根据优化策略着色
                strategy = keyword.get('优化策略', '')
                if '删除' in strategy:
                    table_item.setBackground(QColor(255, 200, 200))
                elif '加价' in strategy or '保持' in strategy:
                    table_item.setBackground(QColor(200, 255, 200))
                elif '降价' in strategy:
                    table_item.setBackground(QColor(255, 255, 200))
                
                self.keywords_table.setItem(row, col, table_item)
    
    def export_data(self):
        """导出所有推广计划的表格为图片"""
        if not self.current_keywords:
            QMessageBox.warning(self, "导出失败", "没有数据可导出")
            return
        
        # 选择导出路径
        folder_path = QFileDialog.getExistingDirectory(
            self, "选择导出路径"
        )
        
        if not folder_path:
            return
        
        try:
            # 创建导出文件夹（使用工具名称）
            export_folder = os.path.join(folder_path, "关键词分析")
            if not os.path.exists(export_folder):
                os.makedirs(export_folder)
            
            # 保存原始列宽
            original_widths = []
            header = self.keywords_table.horizontalHeader()
            for col in range(self.keywords_table.columnCount()):
                original_widths.append(self.keywords_table.columnWidth(col))
            
            # 获取所有推广计划
            plans = set()
            for kw in self.current_keywords:
                plan = kw.get('推广计划', '')
                if plan:
                    plans.add(plan)
            
            if not plans:
                plans = {'未分类'}
            
            # 为每个推广计划导出表格
            for plan_name in sorted(plans):
                # 筛选该推广计划的关键词
                filtered_keywords = [kw for kw in self.current_keywords if kw.get('推广计划', '') == plan_name]
                
                if not filtered_keywords:
                    continue
                
                # 临时更新表格显示该推广计划的数据
                self.update_keywords_table(filtered_keywords)
                
                # 清除表格选中状态
                self.keywords_table.clearSelection()
                
                # 调整列宽以适应内容
                self.keywords_table.resizeColumnsToContents()
                
                # 计算表格的实际宽度（所有列的宽度之和）
                total_width = sum(self.keywords_table.columnWidth(col) for col in range(self.keywords_table.columnCount())) + 1
                
                # 计算表格高度
                header_height = self.keywords_table.horizontalHeader().height()
                row_height = self.keywords_table.rowHeight(0) if self.keywords_table.rowCount() > 0 else 30
                total_height = header_height + (self.keywords_table.rowCount() * row_height) + 1
                
                # 创建 QPixmap 用于绘制（使用实际内容宽度）
                pixmap = QPixmap(total_width, total_height)
                pixmap.fill(QColor(255, 255, 255))  # 白色背景
                
                # 使用 QPainter 绘制表格
                painter = QPainter(pixmap)
                self.keywords_table.render(painter)
                painter.end()
                
                # 生成文件名（替换不合法字符）
                safe_plan_name = plan_name.replace('/', '_').replace('\\', '_').replace(':', '_')
                file_path = os.path.join(export_folder, f"{safe_plan_name}.png")
                
                # 保存图片
                if not pixmap.save(file_path):
                    QMessageBox.critical(self, "导出失败", f"无法保存 {plan_name} 的图片")
                    return
            
            # 恢复原始列宽
            for col in range(self.keywords_table.columnCount()):
                self.keywords_table.setColumnWidth(col, original_widths[col])
            
            # 恢复显示当前推广计划
            if self.plan_combo.count() > 0:
                current_plan = self.plan_combo.currentText()
                filtered_keywords = [kw for kw in self.current_keywords if kw.get('推广计划', '') == current_plan]
                if not filtered_keywords:
                    filtered_keywords = self.current_keywords
                self.update_keywords_table(filtered_keywords)
            
            QMessageBox.information(self, "成功", f"已导出 {len(plans)} 个推广计划的表格到: {export_folder}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出出错: {str(e)}")
    
    def open_settings(self):
        """打开设置对话框"""
        dialog = SettingsDialog(self.config_manager, self, self.recommended_thresholds if hasattr(self, 'recommended_thresholds') else None)
        if dialog.exec_() == QDialog.Accepted:
            # 重新初始化 optimizer 使用新的阈值
            config = self.config_manager.get_thresholds()
            self.optimizer = KeywordOptimizer(config)
            
            # 如果已加载数据，重新分析
            if self.current_keywords:
                # 从 current_keywords 中恢复原始数据（去掉优化建议）
                raw_data = []
                for kw in self.current_keywords:
                    raw_record = {
                        '关键词': kw.get('关键词', ''),
                        '推广计划': kw.get('推广计划', ''),
                        '展现数': kw.get('展现数', 0),
                        '花费': kw.get('花费', 0),
                        '点击率(%)': kw.get('点击率(%)', 0),
                        '转化率(%)': kw.get('转化率(%)', 0),
                        '点击数': kw.get('点击数', 0),
                        '总订单行': kw.get('总订单行', 0),
                        '总订单金额': kw.get('总订单金额', 0),
                    }
                    raw_data.append(raw_record)
                
                # 重新加载数据并分析
                self.optimizer.load_data(raw_data)
                self.current_keywords = self.optimizer.analyze_all_keywords()
                
                # 更新推广计划下拉框
                self.update_plan_combo()
                
                # 直接刷新表格
                if self.plan_combo.count() > 0:
                    plan_name = self.plan_combo.itemText(0)
                    self.on_plan_changed(plan_name)
                
                QMessageBox.information(self, "成功", "阈值已更新，结果已刷新")
    
    def generate_box(self):
        """生成箱唛文件"""
        import pathlib
        
        po_num = self.box_po_input.text().strip()
        warehouse = self.box_warehouse_input.text().strip()
        city_code = self.box_city_input.currentData()
        vendor_name = self.box_vendor_input.text().strip()  # 从文本框获取
        box_seq = self.box_seq_input.text().strip()  # 获取批次
        box_type = self.box_type_input.currentText()  # 获取类型
        
        # 验证输入
        if not all([po_num, warehouse, vendor_name]):
            QMessageBox.warning(self, "警告", "请填写所有必填项")
            return
        
        # 验证采购单号
        if not po_num.isdigit() or len(po_num) != 10:
            QMessageBox.warning(self, "警告", "采购单号必须是10位数字")
            return
        
        # 验证批次
        if not box_seq or not box_seq.isdigit():
            QMessageBox.warning(self, "警告", "批次必须是数字")
            return
        
        try:
            # 创建生成器（使用相对路径）
            generator = CartonMarkGenerator()
            
            # 根据类型确定子文件夹
            if box_type == "3C":
                box_type_folder = "3C箱唛"
            else:  # 玩具
                box_type_folder = "玩具箱唛"
            
            # 生成到templates/箱唛模板/{类型}文件夹
            import pathlib
            gui_dir = pathlib.Path(__file__).resolve().parent
            carton_mark_dir = gui_dir / "templates" / "箱唛模板" / box_type_folder
            carton_mark_dir.mkdir(parents=True, exist_ok=True)
            
            city_label = self.box_city_input.currentText()
            
            # 根据批次生成目标文件名
            if box_seq == "1":
                target_filename = f"{city_code}.{city_label}箱唛.pld"
            else:
                target_filename = f"{city_code}.{city_label}箱唛-{box_seq}.pld"
            target_path = carton_mark_dir / target_filename
            
            # 在生成前检查目标文件是否已存在
            if target_path.exists():
                QMessageBox.warning(self, "文件已存在", f"文件 {target_filename} 已存在，禁止覆盖。\n请使用不同的批次号。")
                return
            
            if generator.generate(po_num, warehouse, city_code, vendor_name, str(carton_mark_dir)):
                # BoxGenerator生成的文件名
                original_filename = f"{city_code}.{city_label}箱唛.pld"
                original_path = carton_mark_dir / original_filename
                
                # 如果需要重命名，则重命名文件
                if original_path != target_path and original_path.exists():
                    import shutil
                    shutil.move(str(original_path), str(target_path))
                
                QMessageBox.information(self, "成功", f"箱唛已生成: {target_path}")
                self.statusBar.showMessage(f"箱唛已生成: {target_path}")
            else:
                QMessageBox.critical(self, "失败", "生成箱唛失败，请检查输入数据")
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[DEBUG] 异常: {error_trace}")
            QMessageBox.critical(self, "错误", f"生成箱唛出错: {str(e)}")
    
    def generate_label(self):
        """生成标签文件"""
        brand = self.label_brand_input.currentText()  # 从下拉栏获取
        product = self.label_product_input.text().strip()
        sku = self.label_sku_input.text().strip()
        label_type = self.label_type_input.currentText()
        
        # 根据类型处理
        if label_type == "螺旋桨":
            # 螺旋桨类型：不需要69码和SN序列号
            code69 = ""
            sn = ""
            sn_seq = "1"  # 默认序号
            propeller_date = self.label_sn_date.text().strip()
            
            if not all([brand, product, sku, propeller_date]):
                QMessageBox.warning(self, "警告", "请填写所有必填项")
                return
            
            # 验证日期
            is_valid, error_msg = self.validate_date(propeller_date)
            if not is_valid:
                QMessageBox.warning(self, "警告", error_msg)
                return
            
            # 验证 SKU 长度
            sku_len = len(sku)
            if sku_len != 12:
                if sku_len < 12:
                    msg = f"SKU 长度不足。当前: {sku_len} 位，需要: 12 位\n\n是否继续生成？"
                else:
                    msg = f"SKU 长度超出。当前: {sku_len} 位，需要: 12 位\n\n是否继续生成？"
                
                reply = QMessageBox.warning(self, "SKU 长度警告", msg, QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.No:
                    return
            
            # 验证产品名称长度（限制 32 字节）
            product_bytes = product.encode('gbk', errors='ignore')
            if len(product_bytes) > 32:
                msg = f"产品名称过长。当前: {len(product_bytes)} 字节，限制: 32 字节\n\n是否继续生成？"
                reply = QMessageBox.warning(self, "产品名称长度警告", msg, QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.No:
                    return
            
            # 弹窗让用户输入文件名
            filename, ok = QInputDialog.getText(
                self, "输入文件名", "请输入螺旋桨标签文件名（不含.pld扩展名）：",
                text=product
            )
            
            if not ok or not filename.strip():
                return
            
            filename = filename.strip()
            
            try:
                # 获取螺旋桨模板路径（使用相对路径）
                import pathlib
                gui_dir = pathlib.Path(__file__).resolve().parent
                template_path = gui_dir / "templates" / "label_propeller.pld"
                
                if not template_path.exists():
                    error_msg = f"生成标签失败\n\n找不到螺旋桨模板文件。请确保模板文件存在于:\n{template_path}"
                    QMessageBox.critical(self, "失败", error_msg)
                    return
                
                print(f"[DEBUG] 使用螺旋桨模板文件: {template_path}")
                
                generator = LabelGenerator(str(template_path))
                generator.set_label_data(brand, product, sku, "", "", propeller_date)
                
                # 生成到templates/标签模板/3C标签文件夹（螺旋桨属于3C）
                label_dir = gui_dir / "templates" / "标签模板" / "3C标签"
                label_dir.mkdir(parents=True, exist_ok=True)
                
                # 检查文件是否已存在
                output_filename = f"{filename}.pld"
                file_path = label_dir / output_filename
                
                if file_path.exists():
                    QMessageBox.warning(self, "文件已存在", f"文件 {output_filename} 已存在，禁止覆盖。\n请使用不同的文件名。")
                    return
                
                # 生成文件
                if generator.generate_pld(str(label_dir), seq_num=sn_seq):
                    # 获取生成的默认文件名
                    default_filename = f"{sn_seq}.{product}.pld"
                    default_path = label_dir / default_filename
                    
                    # 重命名为用户指定的名称
                    if default_path.exists():
                        import shutil
                        shutil.move(str(default_path), str(file_path))
                    
                    QMessageBox.information(self, "成功", f"标签已生成: {file_path}")
                    self.statusBar.showMessage(f"标签已生成: {file_path}")
                else:
                    error_msg = f"生成标签失败\n\n请检查模板文件和输入数据"
                    QMessageBox.critical(self, "失败", error_msg)
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()
                print(f"[DEBUG] 异常: {error_trace}")
                QMessageBox.critical(self, "错误", f"生成标签出错: {str(e)}")
            return
        
        # 3C 和玩具类型的处理
        code69 = self.label_code69_input.text().strip()
        
        # 构建 SN
        sn_prefix = self.label_sn_prefix.currentText()
        sn_seq = self.label_sn_seq.text().strip()
        sn_year = self.label_sn_year.text().strip()  # 从填写框获取
        sn_date = self.label_sn_date.text().strip()
        sn_batch = self.label_sn_batch.text().strip()
        
        if not all([brand, product, sku, code69, sn_seq, sn_year, sn_date, sn_batch]):
            QMessageBox.warning(self, "警告", "请填写所有必填项")
            return
        
        # 验证输入
        if not sn_seq.isdigit() or not (1 <= int(sn_seq) <= 999):
            QMessageBox.warning(self, "警告", "序号必须是 1-999 之间的数字")
            return
        
        # 验证日期
        is_valid, error_msg = self.validate_date(sn_date)
        if not is_valid:
            QMessageBox.warning(self, "警告", error_msg)
            return
        
        if not sn_batch.isdigit() or not (1 <= int(sn_batch) <= 999):
            QMessageBox.warning(self, "警告", "批次号必须是 001-999 之间的数字")
            return
        
        # 验证 SKU 长度
        sku_len = len(sku)
        if sku_len != 12:
            if sku_len < 12:
                msg = f"SKU 长度不足。当前: {sku_len} 位，需要: 12 位\n\n是否继续生成？"
            else:
                msg = f"SKU 长度超出。当前: {sku_len} 位，需要: 12 位\n\n是否继续生成？"
            
            reply = QMessageBox.warning(self, "SKU 长度警告", msg, QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return
        
        # 验证 69码长度
        code69_len = len(code69)
        if code69_len != 13:
            if code69_len < 13:
                msg = f"69码长度不足。当前: {code69_len} 位，需要: 13 位\n\n是否继续生成？"
            else:
                msg = f"69码长度超出。当前: {code69_len} 位，需要: 13 位\n\n是否继续生成？"
            
            reply = QMessageBox.warning(self, "69码长度警告", msg, QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return
        
        # 验证产品名称长度（限制 32 字节）
        product_bytes = product.encode('gbk', errors='ignore')
        if len(product_bytes) > 32:
            msg = f"产品名称过长。当前: {len(product_bytes)} 字节，限制: 32 字节\n\n是否继续生成？"
            reply = QMessageBox.warning(self, "产品名称长度警告", msg, QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return
        
        # 转换年份格式
        converted_year = self.convert_year_format(sn_year, sn_prefix)
        
        # 构建完整 SN
        sn = LabelGenerator.build_sn(sn_prefix, sn_seq, converted_year, sn_date, sn_batch)
        
        try:
            # 获取模板路径（使用相对路径）
            import pathlib
            gui_dir = pathlib.Path(__file__).resolve().parent
            template_path = gui_dir / "templates" / "pld_template.pld"
            
            # 如果默认模板不存在，尝试找任何 .pld 文件
            if not template_path.exists():
                pld_files = list((gui_dir / "templates").glob("label_*.pld"))
                if not pld_files:
                    error_msg = f"生成标签失败\n\n找不到模板文件。请确保模板文件存在于:\n{gui_dir / 'templates'}"
                    QMessageBox.critical(self, "失败", error_msg)
                    return
                template_path = pld_files[0]
            
            print(f"[DEBUG] 使用模板文件: {template_path}")
            
            generator = LabelGenerator(str(template_path))
            generator.set_label_data(brand, product, sku, code69, sn, sn_date)
            
            # 根据SN前缀判断类型
            if sn_prefix == "NB":
                label_type = "玩具标签"
            else:  # SG, SN 等都属于3C
                label_type = "3C标签"
            
            # 生成到templates/标签模板/{类型}文件夹
            label_dir = gui_dir / "templates" / "标签模板" / label_type
            label_dir.mkdir(parents=True, exist_ok=True)
            
            # 检查文件是否已存在
            filename = f"{sn_seq}.{product}.pld"
            file_path = label_dir / filename
            
            if file_path.exists():
                QMessageBox.warning(self, "文件已存在", f"文件 {filename} 已存在，禁止覆盖。\n请使用不同的序号或产品名称。")
                return
            
            if generator.generate_pld(str(label_dir), seq_num=sn_seq):
                QMessageBox.information(self, "成功", f"标签已生成: {file_path}")
                self.statusBar.showMessage(f"标签已生成: {file_path}")
            else:
                error_msg = f"生成标签失败\n\n请检查模板文件和输入数据"
                QMessageBox.critical(self, "失败", error_msg)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[DEBUG] 异常: {error_trace}")
            QMessageBox.critical(self, "错误", f"生成标签出错: {str(e)}")
    
    def closeEvent(self, event):
        """窗口关闭事件"""
        event.accept()
    
    @staticmethod
    def validate_date(date_str: str) -> tuple[bool, str]:
        """
        验证日期格式 MMDD
        
        Args:
            date_str: 日期字符串 (MMDD 格式)
            
        Returns:
            (是否有效, 错误信息)
        """
        if not date_str or len(date_str) != 4 or not date_str.isdigit():
            return False, "日期必须是 MMDD 格式（4位数字）"
        
        month = int(date_str[:2])
        day = int(date_str[2:])
        
        # 验证月份
        if month < 1 or month > 12:
            return False, f"月份必须在 01-12 之间，当前: {month:02d}"
        
        # 每月的天数
        days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        
        # 简单的闰年判断（这里假设当前年份或通用规则）
        # 实际上应该根据具体年份判断，但由于 SN 中的年份可能是 2 位或 4 位
        # 我们使用通用规则：能被 4 整除的是闰年，除非能被 100 整除但不能被 400 整除
        # 为了简化，这里假设 2025, 2026 都不是闰年，但我们可以检查 2 月
        
        max_day = days_in_month[month - 1]
        
        # 特殊处理 2 月（假设可能是闰年）
        if month == 2:
            # 2025 不是闰年，2026 也不是闰年
            # 但为了通用性，我们允许 29 天
            max_day = 29
        
        if day < 1 or day > max_day:
            return False, f"日期 {month:02d} 月的天数必须在 01-{max_day:02d} 之间，当前: {day:02d}"
        
        return True, ""


class RecognizeSettingsDialog(QDialog):
    """识别设置对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_gui = parent
        self.setWindowTitle("识别设置")
        self.setFixedSize(380, 100)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        
        # 初始化配置
        self.recognize_config = self.parent_gui.config_manager.get('recognize_config', {
            'sheet_name': '勿改动！机型',
            'row': 2,
            'column': 10  # J列
        })
        
        # 箱唛识别配置：如果Excel中有"箱唛"工作表，强制使用"箱唛"
        saved_box_config = self.parent_gui.config_manager.get('box_recognize_config', {
            'sheet_name': '箱唛'
        })
        
        if self.parent_gui.current_excel_workbook and "箱唛" in self.parent_gui.current_excel_workbook.sheetnames:
            self.box_recognize_config = {'sheet_name': '箱唛'}
        else:
            self.box_recognize_config = saved_box_config
        
        self.init_ui()
    
    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 第一行：标签识别
        label_layout = QHBoxLayout()
        label_layout.setContentsMargins(0, 0, 0, 0)
        label_layout.setSpacing(5)
        label_layout.addWidget(QLabel("标签识别:"), 0)
        
        # 工作表选择
        label_layout.addWidget(QLabel("工作表:"), 0)
        self.sheet_combo = QComboBox()
        # 获取所有工作表名称
        if self.parent_gui.current_excel_workbook:
            self.sheet_combo.addItems(self.parent_gui.current_excel_workbook.sheetnames)
            # 设置当前值
            if self.recognize_config['sheet_name'] in self.parent_gui.current_excel_workbook.sheetnames:
                self.sheet_combo.setCurrentText(self.recognize_config['sheet_name'])
        self.sheet_combo.setMinimumWidth(120)
        label_layout.addWidget(self.sheet_combo, 0)
        
        # 工作表和单元格之间的间距
        label_layout.addSpacing(10)
        
        # 单元格选择
        label_layout.addWidget(QLabel("单元格:"), 0)
        self.cell_input = QLineEdit()
        # 将行列号转换为单元格序号（如 J2）
        col_letter = chr(64 + self.recognize_config['column'])  # 1->A, 2->B, ..., 10->J
        self.cell_input.setText(f"{col_letter}{self.recognize_config['row']}")
        self.cell_input.setPlaceholderText("如: J2")
        self.cell_input.setMaximumWidth(80)
        label_layout.addWidget(self.cell_input, 0)
        
        label_layout.addStretch()
        layout.addLayout(label_layout)
        
        # 第二行：箱唛识别
        box_layout = QHBoxLayout()
        box_layout.setContentsMargins(0, 0, 0, 0)
        box_layout.setSpacing(5)
        box_layout.addWidget(QLabel("箱唛识别:"), 0)
        
        # 工作表选择
        box_layout.addWidget(QLabel("工作表:"), 0)
        self.box_sheet_combo = QComboBox()
        # 获取所有工作表名称
        if self.parent_gui.current_excel_workbook:
            sheet_names = self.parent_gui.current_excel_workbook.sheetnames
            self.box_sheet_combo.addItems(sheet_names)
            # 强制选择"箱唛"工作表（如果存在，忽略前后空格）
            found_box = False
            for name in sheet_names:
                if name.strip() == "箱唛":
                    self.box_sheet_combo.setCurrentText(name)
                    found_box = True
                    break
            
            if not found_box:
                # 如果"箱唛"不存在，选择保存的配置或第一个工作表
                if self.box_recognize_config['sheet_name'] in sheet_names:
                    self.box_sheet_combo.setCurrentText(self.box_recognize_config['sheet_name'])
                elif sheet_names:
                    self.box_sheet_combo.setCurrentIndex(0)
        self.box_sheet_combo.setMinimumWidth(120)
        box_layout.addWidget(self.box_sheet_combo, 1)
        
        box_layout.addStretch()
        layout.addLayout(box_layout)
        
        layout.addStretch()
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(10)
        
        reset_btn = QPushButton("恢复默认")
        reset_btn.clicked.connect(self.reset_to_default)
        button_layout.addWidget(reset_btn)
        
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def center_on_screen(self):
        """将对话框居中显示在父窗口上"""
        if self.parent_gui:
            parent_geometry = self.parent_gui.geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def accept(self):
        """保存设置并关闭对话框"""
        cell_text = self.cell_input.text().strip().upper()
        
        # 解析单元格序号（如 J2 -> column=10, row=2）
        import re
        match = re.match(r'([A-Z]+)(\d+)', cell_text)
        if match:
            col_letter = match.group(1)
            row_num = int(match.group(2))
            
            # 将列字母转换为数字
            col_num = 0
            for char in col_letter:
                col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            
            self.recognize_config = {
                'sheet_name': self.sheet_combo.currentText(),
                'row': row_num,
                'column': col_num
            }
            
            self.box_recognize_config = {
                'sheet_name': self.box_sheet_combo.currentText()
            }
            
            self.parent_gui.config_manager.set('recognize_config', self.recognize_config)
            self.parent_gui.config_manager.set('box_recognize_config', self.box_recognize_config)
            super().accept()
        else:
            QMessageBox.warning(self, "错误", "单元格格式不正确，请使用如 J2 的格式")
    
    def reset_to_default(self):
        """恢复默认设置"""
        default_config = {
            'sheet_name': '勿改动！机型',
            'row': 2,
            'column': 10  # J列
        }
        default_box_config = {
            'sheet_name': '箱唛'
        }
        self.recognize_config = default_config
        self.box_recognize_config = default_box_config
        self.sheet_combo.setCurrentText(default_config['sheet_name'])
        
        # 恢复箱唛工作表默认值，处理可能带空格的工作表名称
        found_box = False
        if self.parent_gui.current_excel_workbook:
            for name in self.parent_gui.current_excel_workbook.sheetnames:
                if name.strip() == "箱唛":
                    self.box_sheet_combo.setCurrentText(name)
                    found_box = True
                    break
        
        if not found_box:
            # 如果找不到"箱唛"，尝试直接设置
            self.box_sheet_combo.setCurrentText(default_box_config['sheet_name'])
        
        col_letter = chr(64 + default_config['column'])
        self.cell_input.setText(f"{col_letter}{default_config['row']}")


def main():
    app = QApplication(sys.argv)
    window = KeywordAnalyzerGUI()
    window.show()
    sys.exit(app.exec_())


class PropellerFileNameDialog(QDialog):
    """螺旋桨文件命名对话框"""
    
    def __init__(self, parent=None, existing_files=None):
        super().__init__(parent)
        self.existing_files = existing_files or []
        self.filename = None
        self.setWindowTitle("螺旋桨文件命名")
        self.setGeometry(0, 0, 300, 80)
        self.setFixedSize(300, 80)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
    
    def center_on_screen(self):
        """将窗口居中显示"""
        if self.parent():
            parent_geometry = self.parent().geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化 UI"""
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(8)
        
        # 标签和输入框在同一行
        input_layout = QHBoxLayout()
        input_layout.setContentsMargins(0, 0, 0, 0)
        input_layout.setSpacing(5)
        
        label = QLabel("请命名文件:")
        input_layout.addWidget(label)
        
        self.filename_input = QLineEdit()
        self.filename_input.setPlaceholderText("例如: 螺旋桨1")
        input_layout.addWidget(self.filename_input)
        
        layout.addLayout(input_layout)
        
        # 按钮行
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(10)
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept_filename)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        self.filename_input.setFocus()
    
    def accept_filename(self):
        """验证并接受文件名"""
        filename = self.filename_input.text().strip()
        
        if not filename:
            QMessageBox.warning(self, "警告", "文件名不能为空")
            return
        
        # 检查文件名是否已存在
        pld_filename = f"{filename}.pld"
        if pld_filename in self.existing_files:
            QMessageBox.warning(self, "警告", f"文件 {pld_filename} 已存在，请使用其他名称")
            return
        
        self.filename = filename
        self.accept()
    
    def get_filename(self):
        """获取输入的文件名"""
        return self.filename


class LabelDialog(QDialog):
    """标签生成对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_gui = parent
        self.excel_data = None  # 存储导入的Excel数据
        self.setWindowTitle("新增标签")
        self.setGeometry(0, 0, 450, 210)
        self.setFixedSize(450, 210)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
    
    def center_on_screen(self):
        """将窗口居中显示在父窗口上"""
        if self.parent_gui:
            parent_geometry = self.parent_gui.geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化 UI"""
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)
        
        # 输入表单
        form_layout = QGridLayout()
        form_layout.setSpacing(8)
        form_layout.setContentsMargins(0, 0, 0, 0)
        
        # 筛选类型
        form_layout.addWidget(QLabel("类型:"), 0, 0)
        self.label_type_input = QComboBox()
        self.label_type_input.addItems(["3C", "玩具", "螺旋桨"])
        self.label_type_input.currentTextChanged.connect(self.on_type_changed)
        form_layout.addWidget(self.label_type_input, 0, 1)
        
        # 品牌
        form_layout.addWidget(QLabel("品牌:"), 1, 0)
        self.label_brand_input = QComboBox()
        form_layout.addWidget(self.label_brand_input, 1, 1)
        
        # 产品名称
        form_layout.addWidget(QLabel("产品名称:"), 2, 0)
        self.label_product_input = QLineEdit()
        self.label_product_input.setPlaceholderText("例如: JD-sg888+避障")
        form_layout.addWidget(self.label_product_input, 2, 1)
        
        # SKU
        form_layout.addWidget(QLabel("SKU:"), 3, 0)
        self.label_sku_input = QLineEdit()
        self.label_sku_input.setPlaceholderText("例如: 100140692276")
        form_layout.addWidget(self.label_sku_input, 3, 1)
        
        # 69码
        self.label_code69_label = QLabel("69码:")
        form_layout.addWidget(self.label_code69_label, 4, 0)
        self.label_code69_input = QLineEdit()
        self.label_code69_input.setPlaceholderText("例如: 6978030410332")
        form_layout.addWidget(self.label_code69_input, 4, 1)
        
        # SN序列号
        self.label_sn_label = QLabel("SN序列号:")
        form_layout.addWidget(self.label_sn_label, 5, 0)
        
        sn_layout = QHBoxLayout()
        sn_layout.setContentsMargins(0, 0, 0, 0)
        sn_layout.setSpacing(0)  # 改为 0，然后手动添加间距
        
        sn_layout.addWidget(QLabel("类型:"))
        self.label_sn_prefix = QComboBox()
        self.label_sn_prefix.addItems(["SG", "NB"])
        sn_layout.addWidget(self.label_sn_prefix)
        sn_layout.addSpacing(10)
        
        sn_layout.addWidget(QLabel("序号:"))
        self.label_sn_seq = QLineEdit()
        self.label_sn_seq.setPlaceholderText("1-999")
        self.label_sn_seq.setMaximumWidth(50)
        self.label_sn_seq.setMaxLength(3)
        sn_layout.addWidget(self.label_sn_seq)
        sn_layout.addSpacing(10)
        
        sn_layout.addWidget(QLabel("年份:"))
        self.label_sn_year = QLineEdit()
        self.label_sn_year.setPlaceholderText("YYYY")
        self.label_sn_year.setMaximumWidth(60)
        self.label_sn_year.setMaxLength(4)
        from datetime import datetime
        current_year = datetime.now().year
        self.label_sn_year.setText(str(current_year))
        sn_layout.addWidget(self.label_sn_year)
        sn_layout.addSpacing(10)
        
        sn_layout.addWidget(QLabel("日期:"))
        self.label_sn_date = QLineEdit()
        self.label_sn_date.setPlaceholderText("MMDD")
        self.label_sn_date.setMaximumWidth(50)
        self.label_sn_date.setMaxLength(4)
        self.label_sn_date.setText(datetime.now().strftime("%m%d"))
        sn_layout.addWidget(self.label_sn_date)
        
        form_layout.addLayout(sn_layout, 5, 1)
        
        # 批次行（与其他行左对齐）
        form_layout.addWidget(QLabel("批次:"), 6, 0)
        
        batch_layout = QHBoxLayout()
        batch_layout.setContentsMargins(0, 0, 0, 0)
        batch_layout.setSpacing(0)
        
        self.label_sn_batch = QLineEdit()
        self.label_sn_batch.setPlaceholderText("001-999")
        self.label_sn_batch.setMaximumWidth(50)
        self.label_sn_batch.setMaxLength(3)
        self.label_sn_batch.setText("001")
        batch_layout.addWidget(self.label_sn_batch)
        
        batch_layout.addStretch()
        
        export_btn = QPushButton("导出")
        export_btn.clicked.connect(self.export_labels)
        batch_layout.addWidget(export_btn)
        
        form_layout.addLayout(batch_layout, 6, 1)
        
        layout.addLayout(form_layout)
        
        # 初始化品牌列表
        self.on_type_changed("3C")
        
        # 初始化时启用所有表单控件
        self.enable_form_controls()
        
        self.setLayout(layout)
    
    def disable_form_controls(self):
        """禁用所有表单控件"""
        self.label_type_input.setEnabled(False)
        self.label_brand_input.setEnabled(False)
        self.label_product_input.setEnabled(False)
        self.label_sku_input.setEnabled(False)
        self.label_code69_label.setEnabled(False)
        self.label_code69_input.setEnabled(False)
        self.label_sn_label.setEnabled(False)
        self.label_sn_prefix.setEnabled(False)
        self.label_sn_seq.setEnabled(False)
        self.label_sn_year.setEnabled(False)
        self.label_sn_date.setEnabled(False)
        self.label_sn_batch.setEnabled(False)
    
    def enable_form_controls(self):
        """启用所有表单控件"""
        self.label_type_input.setEnabled(True)
        self.label_brand_input.setEnabled(True)
        self.label_product_input.setEnabled(True)
        self.label_sku_input.setEnabled(True)
        self.label_code69_input.setEnabled(True)
        self.label_sn_prefix.setEnabled(True)
        self.label_sn_seq.setEnabled(True)
        self.label_sn_year.setEnabled(True)
        self.label_sn_date.setEnabled(True)
        self.label_sn_batch.setEnabled(True)
    
    def export_labels(self):
        """导出标签到桌面"""
        try:
            import pathlib
            import shutil
            
            # 先生成标签
            label_type = self.label_type_input.currentText()
            if not label_type:
                QMessageBox.warning(self, "警告", "请先选择类型")
                return
            
            # 生成标签并获取生成的文件路径
            generated_file = self.generate_label_and_get_path()
            if not generated_file:
                return
            
            # 目标文件夹（桌面）
            desktop_path = pathlib.Path(os.path.expanduser("~/Desktop"))
            # 螺旋桨标签放在 3C标签 文件夹下
            export_type = "3C标签" if label_type == "螺旋桨" else f"{label_type}标签"
            export_folder = desktop_path / export_type
            export_folder.mkdir(parents=True, exist_ok=True)
            
            # 检查目标文件是否已存在
            dest_file = export_folder / generated_file.name
            if dest_file.exists():
                QMessageBox.warning(
                    self, 
                    "文件已存在", 
                    f"文件 {generated_file.name} 已存在于桌面，禁止覆盖"
                )
                return
            
            # 复制文件到桌面
            shutil.copy2(str(generated_file), str(dest_file))
            
            QMessageBox.information(self, "成功", f"已导出标签到桌面: {export_folder}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def generate_label_and_get_path(self):
        """生成标签并返回生成的文件路径"""
        import pathlib
        
        # 如果导入了Excel，从Excel中读取数据
        if self.excel_data:
            # 从Excel中读取所有行数据（从第2行开始，跳过表头）
            try:
                ws = self.excel_data.active
                generated_files = []
                row_num = 2  # 从第2行开始，跳过表头
                
                # 遍历所有行
                while True:
                    brand = str(ws[f'A{row_num}'].value or "").strip()
                    product = str(ws[f'B{row_num}'].value or "").strip()
                    sku = str(ws[f'C{row_num}'].value or "").strip()
                    code69 = str(ws[f'D{row_num}'].value or "").strip()
                    sn_prefix = str(ws[f'E{row_num}'].value or "SG").strip()
                    sn_seq = str(ws[f'F{row_num}'].value or "").strip()
                    sn_year = str(ws[f'G{row_num}'].value or "").strip()
                    sn_date = str(ws[f'H{row_num}'].value or "").strip()
                    sn_batch = str(ws[f'I{row_num}'].value or "").strip()
                    custom_filename = str(ws[f'J{row_num}'].value or "").strip()  # 文件名列
                    
                    # 如果品牌为空，说明没有更多数据了
                    if not brand:
                        break
                    
                    # 如果69码为空，当作螺旋桨处理
                    if not code69:
                        # 螺旋桨类型
                        if not all([brand, product, sku, sn_date]):
                            QMessageBox.warning(self, "警告", f"第{row_num}行：螺旋桨必填项不完整")
                            return None
                        
                        # 验证日期
                        is_valid, error_msg = self.validate_date(sn_date)
                        if not is_valid:
                            QMessageBox.warning(self, "警告", f"第{row_num}行：{error_msg}")
                            return None
                        
                        try:
                            # 获取螺旋桨模板路径
                            gui_dir = pathlib.Path(__file__).resolve().parent
                            template_path = gui_dir / "templates" / "label_propeller.pld"
                            
                            if not template_path.exists():
                                raise Exception(f"找不到螺旋桨模板文件: {template_path}")
                            
                            generator = LabelGenerator(str(template_path))
                            generator.set_label_data(brand, product, sku, "", "", sn_date)
                            
                            # 生成到templates/标签模板/3C标签文件夹（螺旋桨属于3C）
                            label_dir = gui_dir / "templates" / "标签模板" / "3C标签"
                            label_dir.mkdir(parents=True, exist_ok=True)
                            
                            # 获取现有文件列表
                            existing_files = [f.name for f in label_dir.glob("*.pld")]
                            
                            # 弹窗让用户命名文件
                            dialog = PropellerFileNameDialog(self, existing_files)
                            if dialog.exec_() != QDialog.Accepted:
                                # 用户取消了
                                return None
                            
                            filename = dialog.get_filename()
                            if not filename:
                                return None
                            
                            # 生成文件（使用序号1作为默认）
                            if generator.generate_pld(str(label_dir), seq_num="1"):
                                # 生成的默认文件名
                                default_generated = label_dir / f"1.{product}.pld"
                                
                                # 最终输出文件名
                                output_filename = f"{filename}.pld"
                                output_path = label_dir / output_filename
                                
                                # 重命名生成的文件
                                if default_generated.exists():
                                    default_generated.rename(output_path)
                                
                                if output_path.exists():
                                    generated_files.append(output_path)
                                    
                                    # 添加映射到工作表映射设置
                                    if "螺旋桨" not in self.parent_gui.sheet_mapping:
                                        self.parent_gui.sheet_mapping["螺旋桨"] = {}
                                    
                                    # 使用 SKU 作为映射的 key，文件名作为 value
                                    self.parent_gui.sheet_mapping["螺旋桨"][sku] = filename
                                    self.parent_gui.save_sheet_mapping()
                                    
                                    # 弹窗提醒
                                    QMessageBox.information(self, "成功", f"已添加到螺旋桨映射:\nSKU: {sku}\n文件: {filename}")
                                else:
                                    raise Exception("生成的文件不存在")
                            else:
                                raise Exception("生成标签失败")
                        except Exception as e:
                            QMessageBox.critical(self, "错误", f"第{row_num}行生成螺旋桨标签失败: {str(e)}")
                            return None
                    else:
                        # 普通标签（3C/玩具）
                        if not all([brand, product, sku, code69, sn_prefix, sn_seq, sn_year, sn_date, sn_batch]):
                            QMessageBox.warning(self, "警告", f"第{row_num}行：必填项不完整")
                            return None
                        
                        # 验证输入
                        if not sn_seq.isdigit() or not (1 <= int(sn_seq) <= 999):
                            QMessageBox.warning(self, "警告", f"第{row_num}行：序号必须是1-999之间的数字")
                            return None
                        
                        is_valid, error_msg = self.validate_date(sn_date)
                        if not is_valid:
                            QMessageBox.warning(self, "警告", f"第{row_num}行：{error_msg}")
                            return None
                        
                        if not sn_batch.isdigit() or not (1 <= int(sn_batch) <= 999):
                            QMessageBox.warning(self, "警告", f"第{row_num}行：批次必须是1-999之间的数字")
                            return None
                        
                        try:
                            # 根据前缀转换年份格式
                            year_converted = self.parent_gui.convert_year_format(sn_year, sn_prefix)
                            
                            # 构建完整SN
                            sn = f"{sn_prefix}{sn_seq}{year_converted}{sn_date}{sn_batch}"
                            
                            # 获取模板路径
                            gui_dir = pathlib.Path(__file__).resolve().parent
                            template_path = gui_dir / "templates" / "label_12.pld"
                            
                            if not template_path.exists():
                                raise Exception(f"找不到模板文件: {template_path}")
                            
                            generator = LabelGenerator(str(template_path))
                            generator.set_label_data(brand, product, sku, code69, sn, sn_date)
                            
                            # 根据前缀判断类型
                            if sn_prefix == "NB":
                                label_type_folder = "玩具标签"
                            else:  # SG, SN 等都属于3C
                                label_type_folder = "3C标签"
                            
                            # 生成到templates/标签模板/{type}标签文件夹
                            label_dir = gui_dir / "templates" / "标签模板" / label_type_folder
                            label_dir.mkdir(parents=True, exist_ok=True)
                            
                            # 生成文件
                            if generator.generate_pld(str(label_dir), seq_num=sn_seq):
                                # 确定文件名
                                if custom_filename:
                                    # 使用自定义文件名
                                    output_filename = f"{custom_filename}.pld"
                                else:
                                    # 使用默认命名法
                                    output_filename = f"{sn_seq}.{product}.pld"
                                
                                generated_path = label_dir / output_filename
                                
                                # 如果使用自定义文件名，需要重命名生成的文件
                                if custom_filename:
                                    default_generated = label_dir / f"{sn_seq}.{product}.pld"
                                    if default_generated.exists():
                                        default_generated.rename(generated_path)
                                
                                if generated_path.exists():
                                    generated_files.append(generated_path)
                                else:
                                    raise Exception("生成的文件不存在")
                            else:
                                raise Exception("生成标签失败")
                        except Exception as e:
                            QMessageBox.critical(self, "错误", f"第{row_num}行生成标签失败: {str(e)}")
                            return None
                    
                    row_num += 1
                
                # 返回生成的文件列表
                if generated_files:
                    return generated_files
                else:
                    QMessageBox.warning(self, "警告", "没有有效的数据行")
                    return None
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取Excel数据失败: {str(e)}")
                return None
        else:
            # 从表单中读取数据
            brand = self.label_brand_input.currentText()
            product = self.label_product_input.text().strip()
            sku = self.label_sku_input.text().strip()
            label_type = self.label_type_input.currentText()
            
            # 根据类型处理
            if label_type == "螺旋桨":
                # 螺旋桨类型：不需要69码和SN序列号
                propeller_date = self.label_sn_date.text().strip()
                
                if not all([brand, product, sku, propeller_date]):
                    QMessageBox.warning(self, "警告", "请填写所有必填项")
                    return None
                
                # 验证日期
                is_valid, error_msg = self.validate_date(propeller_date)
                if not is_valid:
                    QMessageBox.warning(self, "警告", error_msg)
                    return None
                
                try:
                    # 获取螺旋桨模板路径
                    gui_dir = pathlib.Path(__file__).resolve().parent
                    template_path = gui_dir / "templates" / "label_propeller.pld"
                    
                    if not template_path.exists():
                        raise Exception(f"找不到螺旋桨模板文件: {template_path}")
                    
                    generator = LabelGenerator(str(template_path))
                    generator.set_label_data(brand, product, sku, "", "", propeller_date)
                    
                    # 生成到templates/标签模板/3C标签文件夹（螺旋桨属于3C）
                    label_dir = gui_dir / "templates" / "标签模板" / "3C标签"
                    label_dir.mkdir(parents=True, exist_ok=True)
                    
                    # 获取现有文件列表
                    existing_files = [f.name for f in label_dir.glob("*.pld")]
                    
                    # 弹窗让用户命名文件
                    dialog = PropellerFileNameDialog(self, existing_files)
                    if dialog.exec_() != QDialog.Accepted:
                        # 用户取消了
                        return None
                    
                    filename = dialog.get_filename()
                    if not filename:
                        return None
                    
                    # 生成文件（使用序号1作为默认）
                    if generator.generate_pld(str(label_dir), seq_num="1"):
                        # 生成的默认文件名
                        default_generated = label_dir / f"1.{product}.pld"
                        
                        # 最终输出文件名
                        output_filename = f"{filename}.pld"
                        output_path = label_dir / output_filename
                        
                        # 重命名生成的文件
                        if default_generated.exists():
                            default_generated.rename(output_path)
                        
                        if output_path.exists():
                            # 添加映射到工作表映射设置
                            if "螺旋桨" not in self.parent_gui.sheet_mapping:
                                self.parent_gui.sheet_mapping["螺旋桨"] = {}
                            
                            # 使用 SKU 作为映射的 key，文件名作为 value
                            self.parent_gui.sheet_mapping["螺旋桨"][sku] = filename
                            self.parent_gui.save_sheet_mapping()
                            
                            # 弹窗提醒
                            QMessageBox.information(self, "成功", f"已添加到螺旋桨映射:\nSKU: {sku}\n文件: {filename}")
                            
                            return output_path
                        else:
                            raise Exception("生成的文件不存在")
                    else:
                        raise Exception("生成标签失败")
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"生成螺旋桨标签失败: {str(e)}")
                    return None
            else:
                # 普通标签（3C/玩具）
                code69 = self.label_code69_input.text().strip()
                sn_prefix = self.label_sn_prefix.currentText()
                sn_seq = self.label_sn_seq.text().strip()
                sn_year = self.label_sn_year.text().strip()
                sn_date = self.label_sn_date.text().strip()
                sn_batch = self.label_sn_batch.text().strip()
                
                if not all([brand, product, sku, code69, sn_prefix, sn_seq, sn_year, sn_date, sn_batch]):
                    QMessageBox.warning(self, "警告", "请填写所有必填项")
                    return None
                
                # 验证输入
                if not sn_seq.isdigit() or not (1 <= int(sn_seq) <= 999):
                    QMessageBox.warning(self, "警告", "序号必须是1-999之间的数字")
                    return None
                
                is_valid, error_msg = self.validate_date(sn_date)
                if not is_valid:
                    QMessageBox.warning(self, "警告", error_msg)
                    return None
                
                if not sn_batch.isdigit() or not (1 <= int(sn_batch) <= 999):
                    QMessageBox.warning(self, "警告", "批次必须是1-999之间的数字")
                    return None
                
                try:
                    # 根据前缀转换年份格式
                    year_converted = self.parent_gui.convert_year_format(sn_year, sn_prefix)
                    
                    # 构建完整SN
                    sn = f"{sn_prefix}{sn_seq}{year_converted}{sn_date}{sn_batch}"
                    
                    # 获取模板路径
                    gui_dir = pathlib.Path(__file__).resolve().parent
                    template_path = gui_dir / "templates" / "label_12.pld"
                    
                    if not template_path.exists():
                        raise Exception(f"找不到模板文件: {template_path}")
                    
                    generator = LabelGenerator(str(template_path))
                    generator.set_label_data(brand, product, sku, code69, sn, sn_date)
                    
                    # 根据前缀判断类型
                    if sn_prefix == "NB":
                        label_type_folder = "玩具标签"
                    else:  # SG, SN 等都属于3C
                        label_type_folder = "3C标签"
                    
                    # 生成到templates/标签模板/{type}标签文件夹
                    label_dir = gui_dir / "templates" / "标签模板" / label_type_folder
                    label_dir.mkdir(parents=True, exist_ok=True)
                    
                    # 生成文件
                    if generator.generate_pld(str(label_dir), seq_num=sn_seq):
                        # 获取生成的文件名
                        generated_filename = f"{sn_seq}.{product}.pld"
                        generated_path = label_dir / generated_filename
                        
                        if generated_path.exists():
                            return generated_path
                        else:
                            raise Exception("生成的文件不存在")
                    else:
                        raise Exception("生成标签失败")
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"生成标签失败: {str(e)}")
                    return None
            try:
                # 根据前缀转换年份格式
                year_converted = self.parent_gui.convert_year_format(sn_year, sn_prefix)
                
                # 构建完整SN
                sn = f"{sn_prefix}{sn_seq}{year_converted}{sn_date}{sn_batch}"
                
                # 获取模板路径
                gui_dir = pathlib.Path(__file__).resolve().parent
                template_path = gui_dir / "templates" / "label_12.pld"
                
                if not template_path.exists():
                    raise Exception(f"找不到模板文件: {template_path}")
                
                generator = LabelGenerator(str(template_path))
                generator.set_label_data(brand, product, sku, code69, sn, sn_date)
                
                # 根据前缀判断类型
                if sn_prefix == "NB":
                    label_type_folder = "玩具标签"
                else:  # SG, SN 等都属于3C
                    label_type_folder = "3C标签"
                
                # 生成到templates/标签模板/{type}标签文件夹
                label_dir = gui_dir / "templates" / "标签模板" / label_type_folder
                label_dir.mkdir(parents=True, exist_ok=True)
                
                # 生成文件
                if generator.generate_pld(str(label_dir), seq_num=sn_seq):
                    # 获取生成的文件名
                    generated_filename = f"{sn_seq}.{product}.pld"
                    generated_path = label_dir / generated_filename
                    
                    if generated_path.exists():
                        return generated_path
                    else:
                        raise Exception("生成的文件不存在")
                else:
                    raise Exception("生成标签失败")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"生成标签失败: {str(e)}")
                return None
    
    
    @staticmethod
    def validate_date(date_str: str) -> tuple:
        """
        验证日期格式 MMDD
        
        Args:
            date_str: 日期字符串 (MMDD 格式)
            
        Returns:
            (是否有效, 错误信息)
        """
        if not date_str or len(date_str) != 4 or not date_str.isdigit():
            return False, "日期必须是 MMDD 格式（4位数字）"
        
        month = int(date_str[:2])
        day = int(date_str[2:])
        
        # 验证月份
        if month < 1 or month > 12:
            return False, f"月份必须在 01-12 之间，当前: {month:02d}"
        
        # 每月的天数
        days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        max_day = days_in_month[month - 1]
        
        # 验证天数
        if day < 1 or day > max_day:
            return False, f"天数必须在 01-{max_day:02d} 之间，当前: {day:02d}"
        
        return True, ""
    
    def generate_propeller_label(self, brand, product, sku, filename=None, propeller_date=None):
        """生成螺旋桨标签"""
        import pathlib
        
        # 如果没有指定文件名，使用产品名称
        if not filename:
            filename = product
        
        # 如果没有指定日期，使用当前日期
        if not propeller_date:
            from datetime import datetime
            propeller_date = datetime.now().strftime("%m%d")
        
        try:
            # 获取螺旋桨模板路径
            gui_dir = pathlib.Path(__file__).resolve().parent
            template_path = gui_dir / "templates" / "label_propeller.pld"
            
            if not template_path.exists():
                raise Exception(f"找不到螺旋桨模板文件: {template_path}")
            
            generator = LabelGenerator(str(template_path))
            generator.set_label_data(brand, product, sku, "", "", propeller_date)
            
            # 生成到templates/标签模板/3C标签文件夹（螺旋桨属于3C）
            label_dir = gui_dir / "templates" / "标签模板" / "3C标签"
            label_dir.mkdir(parents=True, exist_ok=True)
            
            # 检查文件是否已存在
            output_filename = f"{filename}.pld"
            file_path = label_dir / output_filename
            
            if file_path.exists():
                raise Exception(f"文件 {output_filename} 已存在，禁止覆盖")
            
            # 生成文件（使用序号1作为默认）
            if generator.generate_pld(str(label_dir), seq_num="1"):
                # 获取生成的默认文件名
                default_filename = f"1.{product}.pld"
                default_path = label_dir / default_filename
                
                # 重命名为用户指定的名称
                if default_path.exists():
                    import shutil
                    shutil.move(str(default_path), str(file_path))
            else:
                raise Exception("生成标签失败")
        except Exception as e:
            raise Exception(f"生成螺旋桨标签失败: {str(e)}")
    
    def generate_normal_label(self, brand, product, sku, code69, sn, filename=None):
        """生成普通标签（3C/玩具）"""
        import pathlib
        
        # 从SN中解析出序号和前缀
        sn_parts = self.parse_sn(sn)
        if sn_parts:
            seq = sn_parts['seq']
            prefix = sn_parts['prefix']
        else:
            # 解析失败，使用默认序号
            seq = "1"
            prefix = "SG"  # 默认为3C
        
        # 根据前缀判断类型
        if prefix == "NB":
            label_type = "玩具标签"
        else:  # SG, SN 等都属于3C
            label_type = "3C标签"
        
        try:
            # 获取模板路径
            gui_dir = pathlib.Path(__file__).resolve().parent
            template_path = gui_dir / "templates" / "pld_template.pld"
            
            # 如果默认模板不存在，尝试找任何 label_*.pld 文件
            if not template_path.exists():
                pld_files = list((gui_dir / "templates").glob("label_*.pld"))
                if not pld_files:
                    raise Exception(f"找不到模板文件")
                template_path = pld_files[0]
            
            # 获取SN中的日期
            if sn_parts:
                sn_date = sn_parts['date']
            else:
                from datetime import datetime
                sn_date = datetime.now().strftime("%m%d")
            
            generator = LabelGenerator(str(template_path))
            generator.set_label_data(brand, product, sku, code69, sn, sn_date)
            
            # 生成到templates/标签模板/{类型}文件夹
            label_dir = gui_dir / "templates" / "标签模板" / label_type
            label_dir.mkdir(parents=True, exist_ok=True)
            
            # 确定文件名
            if filename:
                # 使用自定义文件名
                output_filename = f"{filename}.pld"
                file_path = label_dir / output_filename
            else:
                # 使用默认格式：序号.产品名称.pld
                output_filename = f"{seq}.{product}.pld"
                file_path = label_dir / output_filename
            
            if file_path.exists():
                raise Exception(f"文件 {output_filename} 已存在，禁止覆盖")
            
            # 生成文件
            if generator.generate_pld(str(label_dir), seq_num=seq):
                # 如果使用自定义文件名，需要重命名
                if filename:
                    default_filename = f"{seq}.{product}.pld"
                    default_path = label_dir / default_filename
                    if default_path.exists():
                        import shutil
                        shutil.move(str(default_path), str(file_path))
            else:
                raise Exception("生成标签失败")
        except Exception as e:
            raise Exception(f"生成普通标签失败: {str(e)}")
    
    def on_type_changed(self, type_name):
        """类型改变时的处理"""
        # 更新品牌列表
        self.label_brand_input.blockSignals(True)
        self.label_brand_input.clear()
        
        if type_name == "螺旋桨":
            # 螺旋桨使用 3C 的品牌列表
            brands = list(self.parent_gui.sheet_mapping["3C"].keys())
            self.label_brand_input.addItems(brands)
        elif type_name in self.parent_gui.sheet_mapping:
            brands = list(self.parent_gui.sheet_mapping[type_name].keys())
            self.label_brand_input.addItems(brands)
        
        self.label_brand_input.blockSignals(False)
        
        # 根据类型设置SN前缀和启用/禁用字段
        if type_name == "3C":
            self.label_sn_prefix.setCurrentText("SG")
            self.label_code69_label.setEnabled(True)
            self.label_code69_input.setEnabled(True)
            self.label_sn_label.setEnabled(True)
            self.label_sn_prefix.setEnabled(True)
            self.label_sn_seq.setEnabled(True)
            self.label_sn_year.setEnabled(True)
            self.label_sn_date.setEnabled(True)
            self.label_sn_batch.setEnabled(True)
        elif type_name == "玩具":
            self.label_sn_prefix.setCurrentText("NB")
            self.label_code69_label.setEnabled(True)
            self.label_code69_input.setEnabled(True)
            self.label_sn_label.setEnabled(True)
            self.label_sn_prefix.setEnabled(True)
            self.label_sn_seq.setEnabled(True)
            self.label_sn_year.setEnabled(True)
            self.label_sn_date.setEnabled(True)
            self.label_sn_batch.setEnabled(True)
        elif type_name == "螺旋桨":
            self.label_code69_label.setEnabled(False)
            self.label_code69_input.setEnabled(False)
            self.label_code69_input.clear()
            self.label_sn_label.setEnabled(False)
            self.label_sn_prefix.setEnabled(False)
            self.label_sn_seq.setEnabled(False)
            self.label_sn_year.setEnabled(False)
            self.label_sn_date.setEnabled(True)
            self.label_sn_batch.setEnabled(False)
    
    def generate_label(self):
        """生成标签"""
        # 调用主窗口的generate_label方法
        self.parent_gui.label_type_input = self.label_type_input
        self.parent_gui.label_brand_input = self.label_brand_input
        self.parent_gui.label_product_input = self.label_product_input
        self.parent_gui.label_sku_input = self.label_sku_input
        self.parent_gui.label_code69_input = self.label_code69_input
        self.parent_gui.label_sn_prefix = self.label_sn_prefix
        self.parent_gui.label_sn_seq = self.label_sn_seq
        self.parent_gui.label_sn_year = self.label_sn_year
        self.parent_gui.label_sn_date = self.label_sn_date
        self.parent_gui.label_sn_batch = self.label_sn_batch
        
        self.parent_gui.generate_label()


class BoxDialog(QDialog):
    """箱唛生成对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_gui = parent
        self.setWindowTitle("新增箱唛")
        self.setGeometry(0, 0, 350, 180)
        self.setFixedSize(350, 180)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
    
    def center_on_screen(self):
        """将窗口居中显示在父窗口上"""
        if self.parent_gui:
            parent_geometry = self.parent_gui.geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化 UI"""
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)
        
        # 箱唛输入表单
        box_form_layout = QGridLayout()
        box_form_layout.setSpacing(8)
        box_form_layout.setContentsMargins(0, 0, 0, 0)
        
        # 筛选类型
        box_form_layout.addWidget(QLabel("类型:"), 0, 0)
        self.box_type_input = QComboBox()
        self.box_type_input.addItems(["3C", "玩具"])
        self.box_type_input.currentTextChanged.connect(self.on_type_changed)
        box_form_layout.addWidget(self.box_type_input, 0, 1)
        
        # 采购单号
        box_form_layout.addWidget(QLabel("采购单号:"), 1, 0)
        self.box_po_input = QLineEdit()
        self.box_po_input.setPlaceholderText("10位数字")
        self.box_po_input.setMaxLength(10)
        box_form_layout.addWidget(self.box_po_input, 1, 1)
        
        # 目的仓
        box_form_layout.addWidget(QLabel("目的仓:"), 2, 0)
        self.box_warehouse_input = QLineEdit()
        self.box_warehouse_input.setPlaceholderText("例如: 仓库A")
        box_form_layout.addWidget(self.box_warehouse_input, 2, 1)
        
        # 目的城市
        box_form_layout.addWidget(QLabel("目的城市:"), 3, 0)
        self.box_city_input = QComboBox()
        city_options = [
            ("1", "北京"),
            ("2", "上海"),
            ("3", "广州"),
            ("4", "成都"),
            ("5", "武汉"),
            ("6", "沈阳"),
            ("7", "西安"),  # 陕西省 西安市
            ("8", "德州")
        ]
        for code, label in city_options:
            self.box_city_input.addItem(label, code)
        box_form_layout.addWidget(self.box_city_input, 3, 1)
        
        # 商家名称
        box_form_layout.addWidget(QLabel("商家名称:"), 4, 0)
        self.box_vendor_input = QLineEdit()
        self.box_vendor_input.setPlaceholderText("根据类型自动填充，可手动修改")
        self.box_vendor_input.setText("ststthmyyx")
        box_form_layout.addWidget(self.box_vendor_input, 4, 1)
        
        # 批次和生成按钮
        box_form_layout.addWidget(QLabel("批次:"), 5, 0)
        batch_button_layout = QHBoxLayout()
        batch_button_layout.setContentsMargins(0, 0, 0, 0)
        batch_button_layout.setSpacing(10)
        
        self.box_seq_input = QLineEdit()
        self.box_seq_input.setPlaceholderText("默认1")
        self.box_seq_input.setText("1")
        self.box_seq_input.setMaximumWidth(60)
        self.box_seq_input.setMaxLength(3)
        batch_button_layout.addWidget(self.box_seq_input)
        
        batch_button_layout.addStretch()
        
        export_btn = QPushButton("导出")
        export_btn.clicked.connect(self.export_box)
        batch_button_layout.addWidget(export_btn)
        
        box_form_layout.addLayout(batch_button_layout, 5, 1)
        
        layout.addLayout(box_form_layout)
        self.setLayout(layout)
    
    def on_type_changed(self, type_name):
        """类型改变时的处理"""
        if type_name == "3C":
            self.box_vendor_input.setText("ststthmyyx")
        elif type_name == "玩具":
            self.box_vendor_input.setText("stsnb")
    
    def export_box(self):
        """导出箱唛到桌面"""
        try:
            import pathlib
            import shutil
            
            box_type = self.box_type_input.currentText()
            if not box_type:
                QMessageBox.warning(self, "警告", "请先选择类型")
                return
            
            # 生成箱唛并获取生成的文件路径
            generated_file = self.generate_box_and_get_path()
            if not generated_file:
                return
            
            # 目标文件夹（桌面）
            desktop_path = pathlib.Path(os.path.expanduser("~/Desktop"))
            export_folder = desktop_path / f"{box_type}箱唛"
            export_folder.mkdir(parents=True, exist_ok=True)
            
            # 检查目标文件是否已存在
            dest_file = export_folder / generated_file.name
            if dest_file.exists():
                QMessageBox.warning(
                    self, 
                    "文件已存在", 
                    f"文件 {generated_file.name} 已存在于桌面，禁止覆盖"
                )
                return
            
            # 复制文件到桌面
            shutil.copy2(str(generated_file), str(dest_file))
            
            QMessageBox.information(self, "成功", f"已导出箱唛到桌面: {export_folder}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def generate_box_and_get_path(self):
        """生成箱唛并返回生成的文件路径"""
        import pathlib
        
        po_num = self.box_po_input.text().strip()
        warehouse = self.box_warehouse_input.text().strip()
        city_code = self.box_city_input.currentData()
        vendor_name = self.box_vendor_input.text().strip()
        box_seq = self.box_seq_input.text().strip()
        box_type = self.box_type_input.currentText()
        
        # 验证输入
        if not all([po_num, warehouse, vendor_name]):
            QMessageBox.warning(self, "警告", "请填写所有必填项")
            return None
        
        # 验证采购单号
        if not po_num.isdigit() or len(po_num) != 10:
            QMessageBox.warning(self, "警告", "采购单号必须是10位数字")
            return None
        
        # 验证批次
        if not box_seq or not box_seq.isdigit():
            QMessageBox.warning(self, "警告", "批次必须是数字")
            return None
        
        try:
            import shutil
            
            # 根据类型确定子文件夹
            if box_type == "3C":
                box_type_folder = "3C箱唛"
            else:  # 玩具
                box_type_folder = "玩具箱唛"
            
            # 生成到templates/箱唛模板/{类型}文件夹
            gui_dir = pathlib.Path(__file__).resolve().parent
            carton_mark_dir = gui_dir / "templates" / "箱唛模板" / box_type_folder
            carton_mark_dir.mkdir(parents=True, exist_ok=True)
            
            city_label = self.box_city_input.currentText()
            
            # 根据批次生成目标文件名
            if box_seq == "1":
                target_filename = f"{city_code}.{city_label}箱唛.pld"
            else:
                target_filename = f"{city_code}.{city_label}箱唛-{box_seq}.pld"
            target_path = carton_mark_dir / target_filename
            
            # 在生成前检查目标文件是否已存在
            if target_path.exists():
                QMessageBox.warning(self, "文件已存在", f"文件 {target_filename} 已存在，禁止覆盖。\n请使用不同的批次号。")
                return None
            
            # 创建生成器并生成箱唛
            generator = CartonMarkGenerator()
            
            if generator.generate(po_num, warehouse, city_code, vendor_name, str(carton_mark_dir)):
                # 生成器生成的文件名
                original_filename = f"{city_code}.{city_label}箱唛.pld"
                original_path = carton_mark_dir / original_filename
                
                # 如果需要重命名（批次不是1），则重命名文件
                if original_path != target_path and original_path.exists():
                    shutil.move(str(original_path), str(target_path))
                
                return target_path
            else:
                QMessageBox.critical(self, "失败", "生成箱唛失败，请检查输入数据")
                return None
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[DEBUG] 异常: {error_trace}")
            QMessageBox.critical(self, "错误", f"生成箱唛出错: {str(e)}")
            return None
    
    def generate_box(self):
        """生成箱唛（添加模板按钮）"""
        generated_file = self.generate_box_and_get_path()
        if generated_file:
            QMessageBox.information(self, "成功", f"箱唛已生成: {generated_file.name}")
    
    @staticmethod
    def validate_date(date_str: str) -> tuple:
        """
        验证日期格式 MMDD
        
        Args:
            date_str: 日期字符串 (MMDD 格式)
            
        Returns:
            (是否有效, 错误信息)
        """
        if not date_str or len(date_str) != 4 or not date_str.isdigit():
            return False, "日期必须是 MMDD 格式（4位数字）"
        
        month = int(date_str[:2])
        day = int(date_str[2:])
        
        # 验证月份
        if month < 1 or month > 12:
            return False, f"月份必须在 01-12 之间，当前: {month:02d}"
        
        # 每月的天数
        days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        max_day = days_in_month[month - 1]
        
        # 验证天数
        if day < 1 or day > max_day:
            return False, f"天数必须在 01-{max_day:02d} 之间，当前: {day:02d}"
        
        return True, ""


class SheetMappingDialog(QDialog):
    """工作表映射设置对话框"""
    
    def __init__(self, parent=None, mapping=None):
        super().__init__(parent)
        self.parent_gui = parent
        import copy
        self.mapping = copy.deepcopy(mapping) if mapping else {}
        
        # 如果mapping为空，使用默认值
        if not self.mapping and parent:
            self.mapping = parent.init_sheet_mapping()
        
        # 加载已保存的规则
        self.rules = {}
        if parent and hasattr(parent, 'sheet_rules'):
            self.rules = copy.deepcopy(parent.sheet_rules)
        
        # 全局锁定状态，每次打开都是锁定状态
        self.is_locked = True
        
        self.setWindowTitle("工作表映射设置")
        self.setGeometry(0, 0, 400, 300)
        self.setFixedSize(400, 260)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
        self.update_lock_state()
    
    def center_on_screen(self):
        """将窗口居中显示在父窗口上"""
        if self.parent_gui:
            parent_geometry = self.parent_gui.geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # 类型选择和锁定按钮
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("类型:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems(["3C", "玩具", "螺旋桨"])
        self.type_combo.currentTextChanged.connect(self.on_type_changed)
        type_layout.addWidget(self.type_combo)
        type_layout.addStretch()
        
        # 锁定/解锁按钮
        self.lock_btn = QPushButton("解锁")
        self.lock_btn.setMaximumWidth(60)
        self.lock_btn.clicked.connect(self.toggle_lock)
        type_layout.addWidget(self.lock_btn)
        
        layout.addLayout(type_layout)
        
        # 品牌和工作表映射表
        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(3)
        self.mapping_table.setHorizontalHeaderLabels(["品牌/商品编号", "工作表/模板文件名", "规则"])
        self.mapping_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.mapping_table.setColumnWidth(0, 100)
        self.mapping_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.mapping_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.mapping_table.setColumnWidth(2, 120)
        # 隐藏行号（序号列）
        self.mapping_table.verticalHeader().setVisible(False)
        # 禁止编辑
        self.mapping_table.setEditTriggers(QTableWidget.NoEditTriggers)
        # 允许选择单元格
        self.mapping_table.setSelectionBehavior(QTableWidget.SelectItems)
        self.mapping_table.itemClicked.connect(self.on_table_item_clicked)
        self.mapping_table.itemDoubleClicked.connect(self.on_table_item_double_clicked)
        layout.addWidget(self.mapping_table)
        
        # 按钮行
        button_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("新增")
        self.add_btn.clicked.connect(self.add_mapping)
        button_layout.addWidget(self.add_btn)
        
        self.delete_btn = QPushButton("删除")
        self.delete_btn.clicked.connect(self.delete_mapping)
        button_layout.addWidget(self.delete_btn)
        
        button_layout.addStretch()
        
        self.reset_btn = QPushButton("恢复默认")
        self.reset_btn.clicked.connect(self.reset_defaults)
        button_layout.addWidget(self.reset_btn)
        
        self.ok_btn = QPushButton("确定")
        self.ok_btn.clicked.connect(self.save_and_close)
        button_layout.addWidget(self.ok_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # 初始化显示
        self.on_type_changed("3C")
    
    def on_type_changed(self, type_name):
        """类型改变时更新表格"""
        try:
            self.mapping_table.setRowCount(0)
            
            # 根据类型禁用或启用"恢复默认"按钮
            if type_name == "螺旋桨":
                self.reset_btn.setEnabled(False)
            else:
                self.reset_btn.setEnabled(True)
            
            if not self.mapping or type_name not in self.mapping:
                return
            
            row = 0
            for key, value in self.mapping[type_name].items():
                # 对于螺旋桨，key 是商品编号，value 是模板文件名
                # 对于 3C 和玩具，key 是品牌，value 是工作表列表
                
                if type_name == "螺旋桨":
                    # 螺旋桨映射：商品编号 -> 模板文件名
                    self.mapping_table.insertRow(row)
                    
                    # 商品编号列
                    key_item = QTableWidgetItem(key)
                    key_item.setFlags(key_item.flags() & ~Qt.ItemIsEditable)
                    self.mapping_table.setItem(row, 0, key_item)
                    
                    # 模板文件名列
                    value_item = QTableWidgetItem(value)
                    value_item.setFlags(value_item.flags() & ~Qt.ItemIsEditable)
                    
                    self.mapping_table.setItem(row, 1, value_item)
                    
                    # 规则列下拉框（螺旋桨暂无规则，设置为禁用）
                    rule_combo = QComboBox()
                    rule_combo.addItem("无规则")
                    rule_combo.setEnabled(False)  # 禁用下拉框，显示为灰色
                    self.mapping_table.setCellWidget(row, 2, rule_combo)
                    
                    row += 1
                else:
                    # 3C 和玩具映射：品牌 -> 工作表列表
                    sheets = value
                    if isinstance(sheets, list):
                        for sheet in sheets:
                            self.mapping_table.insertRow(row)
                            
                            # 品牌列（每行都显示品牌名）
                            brand_item = QTableWidgetItem(key)
                            brand_item.setFlags(brand_item.flags() & ~Qt.ItemIsEditable)
                            self.mapping_table.setItem(row, 0, brand_item)
                            
                            # 工作表列
                            sheet_item = QTableWidgetItem(sheet)
                            sheet_item.setFlags(sheet_item.flags() & ~Qt.ItemIsEditable)
                            self.mapping_table.setItem(row, 1, sheet_item)
                            
                            # 规则列下拉框
                            rule_combo = QComboBox()
                            rule_combo.addItem("无规则")
                            rule_combo.addItem("跳过红色文字行")
                            
                            # 从已保存的规则中获取
                            rule_key = f"{key}|{sheet}"
                            if type_name in self.rules and rule_key in self.rules[type_name]:
                                rule_code = self.rules[type_name][rule_key]
                                # 将规则代码转换为中文名称
                                if rule_code == "skip_red_text":
                                    rule_combo.setCurrentText("跳过红色文字行")
                            
                            self.mapping_table.setCellWidget(row, 2, rule_combo)
                            
                            row += 1
                    else:
                        # 如果不是列表，当作单个工作表处理
                        self.mapping_table.insertRow(row)
                        brand_item = QTableWidgetItem(key)
                        brand_item.setFlags(brand_item.flags() & ~Qt.ItemIsEditable)
                        self.mapping_table.setItem(row, 0, brand_item)
                        sheet_item = QTableWidgetItem(sheets)
                        sheet_item.setFlags(sheet_item.flags() & ~Qt.ItemIsEditable)
                        self.mapping_table.setItem(row, 1, sheet_item)
                        
                        # 规则列下拉框
                        rule_combo = QComboBox()
                        rule_combo.addItem("无规则")
                        rule_combo.addItem("跳过红色文字行")
                        
                        # 从已保存的规则中获取
                        rule_key = f"{key}|{sheets}"
                        if type_name in self.rules and rule_key in self.rules[type_name]:
                            rule_code = self.rules[type_name][rule_key]
                            # 将规则代码转换为中文名称
                            if rule_code == "skip_red_text":
                                rule_combo.setCurrentText("跳过红色文字行")
                        
                        self.mapping_table.setCellWidget(row, 2, rule_combo)
                        
                        row += 1
        except Exception as e:
            print(f"[DEBUG] on_type_changed error: {str(e)}")
    
    def add_mapping(self):
        """新增映射"""
        type_name = self.type_combo.currentText()
        
        if type_name == "螺旋桨":
            # 螺旋桨映射：商品编号 -> 模板文件名
            sku, ok = self.show_input_dialog("新增商品编号", "请输入商品编号:")
            if not ok or not sku.strip():
                return
            
            sku = sku.strip()
            
            template_file, ok = self.show_input_dialog("新增模板文件名", "请输入模板文件名:")
            if not ok or not template_file.strip():
                return
            
            template_file = template_file.strip()
            
            if type_name not in self.mapping:
                self.mapping[type_name] = {}
            
            self.mapping[type_name][sku] = template_file
        else:
            # 3C 和玩具映射：品牌 -> 工作表列表
            brand, ok = self.show_input_dialog("新增品牌", "请输入品牌名称:")
            if not ok or not brand.strip():
                return
            
            brand = brand.strip()
            
            sheets_text, ok = self.show_input_dialog("新增工作表", "请输入工作表名称(多个用逗号分隔):")
            if not ok or not sheets_text.strip():
                return
            
            sheets = [s.strip() for s in sheets_text.split(",")]
            
            if type_name not in self.mapping:
                self.mapping[type_name] = {}
            
            self.mapping[type_name][brand] = sheets
        
        self.on_type_changed(type_name)
    
    def show_input_dialog(self, title, label_text, default_text=""):
        """显示自定义输入对话框，只有保存按钮"""
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowType.WindowContextHelpButtonHint)
        dialog.setModal(True)
        dialog.setMinimumWidth(350)
        
        layout = QVBoxLayout()
        
        # 标签
        label = QLabel(label_text)
        layout.addWidget(label)
        
        # 输入框和保存按钮在同一行
        input_layout = QHBoxLayout()
        text_input = QLineEdit()
        text_input.setText(default_text)
        input_layout.addWidget(text_input)
        
        save_button = QPushButton("保存")
        save_button.setMaximumWidth(60)
        save_button.clicked.connect(dialog.accept)
        input_layout.addWidget(save_button)
        
        layout.addLayout(input_layout)
        dialog.setLayout(layout)
        
        ok = dialog.exec() == QDialog.DialogCode.Accepted
        return text_input.text(), ok
    
    def delete_mapping(self):
        """删除映射"""
        # 如果被锁定，不允许删除
        if self.is_locked:
            return
        
        type_name = self.type_combo.currentText()
        current_row = self.mapping_table.currentRow()
        
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请先选择要删除的行")
            return
        
        # 获取当前行的第一列和第二列
        col0_item = self.mapping_table.item(current_row, 0)
        col1_item = self.mapping_table.item(current_row, 1)
        
        if not col0_item or not col1_item:
            return
        
        col0_text = col0_item.text()
        col1_text = col1_item.text()
        
        if not col0_text or type_name not in self.mapping or col0_text not in self.mapping[type_name]:
            return
        
        # 弹窗确认删除
        reply = QMessageBox.question(self, "确认删除", f"确定要删除 {col0_text} 的映射吗?")
        if reply != QMessageBox.Yes:
            return
        
        if type_name == "螺旋桨":
            # 螺旋桨映射：直接删除商品编号
            del self.mapping[type_name][col0_text]
        else:
            # 3C 和玩具映射：从品牌的工作表列表中删除该工作表
            sheets = self.mapping[type_name][col0_text]
            if isinstance(sheets, list) and col1_text in sheets:
                sheets.remove(col1_text)
                
                # 如果该品牌没有工作表了，删除整个品牌
                if not sheets:
                    del self.mapping[type_name][col0_text]
        
        # 保存映射
        self.parent_gui.sheet_mapping = self.mapping
        self.parent_gui.save_sheet_mapping()
        
        self.on_type_changed(type_name)
    
    def reset_defaults(self):
        """恢复默认设置"""
        reply = QMessageBox.question(self, "确认", "确定要恢复默认设置吗?")
        if reply == QMessageBox.Yes:
            self.mapping = self.parent_gui.init_sheet_mapping()
            self.on_type_changed(self.type_combo.currentText())
    
    def on_table_item_clicked(self, item):
        """表格项点击时的处理"""
        # 这个方法现在不需要做任何事，因为下拉框已经在表格中
        pass
    
    def on_table_item_double_clicked(self, item):
        """表格项双击时的处理 - 编辑映射"""
        # 如果被锁定，不允许编辑
        if self.is_locked:
            return
        
        row = item.row()
        col = item.column()
        type_name = self.type_combo.currentText()
        
        # 获取当前行的品牌/商品编号和工作表/模板文件名
        col0_item = self.mapping_table.item(row, 0)
        col1_item = self.mapping_table.item(row, 1)
        
        if not col0_item or not col1_item:
            return
        
        old_key = col0_item.text()
        old_value = col1_item.text()
        
        if type_name == "螺旋桨":
            # 螺旋桨映射：编辑商品编号或模板文件名
            if col == 0:
                # 编辑商品编号
                new_key, ok = self.show_input_dialog("编辑商品编号", "请输入新的商品编号:", old_key)
                if ok and new_key.strip() and new_key != old_key:
                    new_key = new_key.strip()
                    if new_key in self.mapping[type_name]:
                        QMessageBox.warning(self, "警告", "该商品编号已存在")
                        return
                    # 更新映射
                    self.mapping[type_name][new_key] = self.mapping[type_name].pop(old_key)
                    self.on_type_changed(type_name)
            elif col == 1:
                # 编辑模板文件名
                new_value, ok = self.show_input_dialog("编辑模板文件名", "请输入新的模板文件名:", old_value)
                if ok and new_value.strip() and new_value != old_value:
                    new_value = new_value.strip()
                    self.mapping[type_name][old_key] = new_value
                    self.on_type_changed(type_name)
        else:
            # 3C 和玩具映射：编辑品牌或工作表
            if col == 0:
                # 编辑品牌
                new_key, ok = self.show_input_dialog("编辑品牌", "请输入新的品牌名称:", old_key)
                if ok and new_key.strip() and new_key != old_key:
                    new_key = new_key.strip()
                    if new_key in self.mapping[type_name]:
                        QMessageBox.warning(self, "警告", "该品牌已存在")
                        return
                    # 更新映射
                    self.mapping[type_name][new_key] = self.mapping[type_name].pop(old_key)
                    self.on_type_changed(type_name)
            elif col == 1:
                # 编辑工作表
                new_value, ok = self.show_input_dialog("编辑工作表", "请输入新的工作表名称:", old_value)
                if ok and new_value.strip() and new_value != old_value:
                    new_value = new_value.strip()
                    # 从旧工作表列表中删除，添加新工作表
                    sheets = self.mapping[type_name][old_key]
                    if isinstance(sheets, list):
                        if old_value in sheets:
                            sheets.remove(old_value)
                        if new_value not in sheets:
                            sheets.append(new_value)
                    self.on_type_changed(type_name)
    
    def toggle_lock(self):
        """切换全局锁定状态"""
        self.is_locked = not self.is_locked
        self.update_lock_state()
    
    def update_lock_state(self):
        """更新锁定状态下的UI"""
        # 更新按钮文本和状态
        if self.is_locked:
            self.lock_btn.setText("解锁")
        else:
            self.lock_btn.setText("锁定")
        
        # 禁用/启用所有控件
        self.type_combo.setEnabled(not self.is_locked)
        self.mapping_table.setEnabled(not self.is_locked)
        self.add_btn.setEnabled(not self.is_locked)
        self.delete_btn.setEnabled(not self.is_locked)
        self.reset_btn.setEnabled(not self.is_locked)
        self.ok_btn.setEnabled(not self.is_locked)
        
        # 设置表格单元格的背景色
        if self.is_locked:
            # 锁定时，所有单元格变灰色
            for row in range(self.mapping_table.rowCount()):
                for col in range(self.mapping_table.columnCount()):
                    item = self.mapping_table.item(row, col)
                    if item:
                        item.setBackground(QColor(200, 200, 200))
                    # 下拉框也要变灰色
                    widget = self.mapping_table.cellWidget(row, col)
                    if widget:
                        widget.setEnabled(False)
        else:
            # 解锁时，恢复正常颜色
            for row in range(self.mapping_table.rowCount()):
                for col in range(self.mapping_table.columnCount()):
                    item = self.mapping_table.item(row, col)
                    if item:
                        item.setBackground(QColor(255, 255, 255))
                    # 下拉框恢复可用
                    widget = self.mapping_table.cellWidget(row, col)
                    if widget:
                        widget.setEnabled(True)
    
    def save_and_close(self):
        """保存并关闭"""
        # 从表格中提取规则信息
        self.extract_rules_from_table()
        self.accept()
    
    def extract_rules_from_table(self):
        """从表格中提取规则信息"""
        # 创建规则存储结构
        if not hasattr(self, 'rules'):
            self.rules = {}
        
        type_name = self.type_combo.currentText()
        if type_name not in self.rules:
            self.rules[type_name] = {}
        
        # 遍历表格中的所有行
        for row in range(self.mapping_table.rowCount()):
            col0_item = self.mapping_table.item(row, 0)
            col1_item = self.mapping_table.item(row, 1)
            
            if not col0_item or not col1_item:
                continue
            
            brand = col0_item.text()
            sheet = col1_item.text()
            
            # 从下拉框中获取规则
            rule_combo = self.mapping_table.cellWidget(row, 2)
            if rule_combo and isinstance(rule_combo, QComboBox):
                rule_display = rule_combo.currentText()
                
                # 将中文名称转换回规则代码
                rule_code = ""
                if rule_display == "跳过红色文字行":
                    rule_code = "skip_red_text"
                
                # 创建规则键（品牌-工作表）
                rule_key = f"{brand}|{sheet}"
                self.rules[type_name][rule_key] = rule_code
    
    def get_mapping(self):
        """获取映射关系"""
        return self.mapping
    
    def get_rules(self):
        """获取规则关系"""
        if not hasattr(self, 'rules'):
            self.rules = {}
        return self.rules
        """获取映射关系"""
        return self.mapping


class OrderFormDeliveryDialog(QDialog):
    """预定表预计到货日期设置对话框"""
    
    def __init__(self, cities, parent=None):
        super().__init__(parent)
        self.cities = cities
        self.delivery_days = {}
        self.setWindowTitle("设置预计到货日期")
        self.setGeometry(0, 0, 400, 300)
        self.setFixedSize(400, 300)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
    
    def center_on_screen(self):
        """将窗口居中显示在父窗口上"""
        if self.parent():
            parent_geometry = self.parent().geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化 UI"""
        main_layout = QVBoxLayout()
        
        # 说明文字
        info_label = QLabel("请为每个城市设置预计到货日期（当前时间+N天）:")
        main_layout.addWidget(info_label)
        
        # 城市输入区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        
        self.city_spinboxes = {}
        
        for city in self.cities:
            city_layout = QHBoxLayout()
            city_layout.setSpacing(10)
            
            city_label = QLabel(city)
            city_label.setMinimumWidth(80)
            city_layout.addWidget(city_label)
            
            spinbox = QSpinBox()
            spinbox.setMinimum(0)
            spinbox.setMaximum(30)
            spinbox.setValue(3)  # 默认3天
            spinbox.setSuffix(" 天")
            self.city_spinboxes[city] = spinbox
            city_layout.addWidget(spinbox)
            
            city_layout.addStretch()
            scroll_layout.addLayout(city_layout)
        
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        main_layout.addWidget(scroll)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.save_delivery_days)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
    
    def save_delivery_days(self):
        """保存预计到货日期"""
        for city, spinbox in self.city_spinboxes.items():
            self.delivery_days[city] = spinbox.value()
        self.accept()
    
    def get_delivery_days(self):
        """获取预计到货日期"""
        return self.delivery_days


class SettingsDialog(QDialog):
    """设置对话框"""
    
    def __init__(self, config_manager, parent=None, recommended=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.recommended = recommended or {}
        self.setWindowTitle("阈值设置")
        self.setGeometry(0, 0, 300, 140)
        self.setFixedSize(300, 140)
        # 移除问号按钮
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
    
    def center_on_screen(self):
        """将窗口居中显示在父窗口上"""
        if self.parent():
            parent_geometry = self.parent().geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化 UI"""
        main_layout = QVBoxLayout()
        
        # 获取当前配置
        config = self.config_manager.get_thresholds()
        
        # 设置项布局
        settings_layout = QGridLayout()
        settings_layout.setSpacing(5)
        
        row = 0
        
        # 展现数阈值
        settings_layout.addWidget(QLabel("展现数:"), row, 0)
        self.impression_spin = QSpinBox()
        self.impression_spin.setMaximum(999999)
        self.impression_spin.setValue(config['impression_threshold'])
        settings_layout.addWidget(self.impression_spin, row, 1)
        if self.recommended:
            rec_text = f"推荐值: {self.recommended.get('impression_threshold', 100)}"
            settings_layout.addWidget(QLabel(rec_text), row, 2)
        row += 1
        
        # 花费阈值
        settings_layout.addWidget(QLabel("花费:"), row, 0)
        self.cost_spin = QDoubleSpinBox()
        self.cost_spin.setMaximum(999999)
        self.cost_spin.setValue(config['cost_threshold'])
        settings_layout.addWidget(self.cost_spin, row, 1)
        if self.recommended:
            rec_text = f"推荐值: {self.recommended.get('cost_threshold', 50):.2f}"
            settings_layout.addWidget(QLabel(rec_text), row, 2)
        row += 1
        
        # 点击率阈值
        settings_layout.addWidget(QLabel("点击率 (%):"), row, 0)
        self.ctr_spin = QDoubleSpinBox()
        self.ctr_spin.setMaximum(100)
        self.ctr_spin.setDecimals(2)
        self.ctr_spin.setValue(config['ctr_threshold'])
        settings_layout.addWidget(self.ctr_spin, row, 1)
        if self.recommended:
            rec_text = f"推荐值: {self.recommended.get('ctr_threshold', 3.0):.2f}%"
            settings_layout.addWidget(QLabel(rec_text), row, 2)
        row += 1
        
        # 转化率阈值
        settings_layout.addWidget(QLabel("转化率 (%):"), row, 0)
        self.conversion_spin = QDoubleSpinBox()
        self.conversion_spin.setMaximum(100)
        self.conversion_spin.setDecimals(2)
        self.conversion_spin.setValue(config['conversion_threshold'])
        settings_layout.addWidget(self.conversion_spin, row, 1)
        if self.recommended:
            rec_text = f"推荐值: {self.recommended.get('conversion_threshold', 1.0):.2f}%"
            settings_layout.addWidget(QLabel(rec_text), row, 2)
        
        # 设置第2列伸展
        settings_layout.setColumnStretch(1, 1)
        
        main_layout.addLayout(settings_layout)
        main_layout.addStretch()
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        reset_btn = QPushButton("恢复默认")
        reset_btn.clicked.connect(self.reset_defaults)
        button_layout.addWidget(reset_btn)
        
        if self.recommended:
            use_rec_btn = QPushButton("使用推荐值")
            use_rec_btn.clicked.connect(self.use_recommended)
            button_layout.addWidget(use_rec_btn)
        
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.save_settings)
        button_layout.addWidget(ok_btn)
        
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
    
    def reset_defaults(self):
        """恢复默认值"""
        from main.common.config import ConfigManager
        defaults = ConfigManager.DEFAULT_CONFIG
        self.impression_spin.setValue(defaults['impression_threshold'])
        self.cost_spin.setValue(defaults['cost_threshold'])
        self.ctr_spin.setValue(defaults['ctr_threshold'])
        self.conversion_spin.setValue(defaults['conversion_threshold'])
    
    def use_recommended(self):
        """使用推荐值"""
        if self.recommended:
            self.impression_spin.setValue(self.recommended.get('impression_threshold', 100))
            self.cost_spin.setValue(self.recommended.get('cost_threshold', 50))
            self.ctr_spin.setValue(self.recommended.get('ctr_threshold', 3.0))
            self.conversion_spin.setValue(self.recommended.get('conversion_threshold', 1.0))
    
    def save_settings(self):
        """保存设置"""
        thresholds = {
            'impression_threshold': self.impression_spin.value(),
            'cost_threshold': self.cost_spin.value(),
            'ctr_threshold': self.ctr_spin.value(),
            'conversion_threshold': self.conversion_spin.value(),
        }
        
        if self.config_manager.set_thresholds(thresholds):
            QMessageBox.information(self, "成功", "阈值已保存")
            self.accept()
        else:
            QMessageBox.critical(self, "失败", "保存阈值失败")


class AppSettingsDialog(QDialog):
    """全局设置对话框"""
    
    def __init__(self, parent=None, app_version="1.0", export_path=""):
        super().__init__(parent)
        self.app_version = app_version
        self.export_path = export_path.replace('/', '\\')  # 统一转换为反斜杠
        self.release_date = "未知"  # 发布日期，默认为未知
        self.setWindowTitle("全局设置")
        self.setGeometry(0, 0, 380, 170)
        self.setFixedSize(380, 170)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.center_on_screen()
        self.init_ui()
    
    def center_on_screen(self):
        """将窗口居中显示在父窗口上"""
        if self.parent():
            parent_geometry = self.parent().geometry()
            x = parent_geometry.x() + (parent_geometry.width() - self.width()) // 2
            y = parent_geometry.y() + (parent_geometry.height() - self.height()) // 2
            self.move(x, y)
        else:
            screen = self.screen()
            screen_geometry = screen.geometry()
            x = (screen_geometry.width() - self.width()) // 2
            y = (screen_geometry.height() - self.height()) // 2
            self.move(x, y)
    
    def init_ui(self):
        """初始化 UI"""
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # 版本信息
        version_group = QGroupBox("版本信息")
        version_layout = QVBoxLayout()
        version_layout.setSpacing(5)
        
        version_info_layout = QHBoxLayout()
        version_info_layout.addStretch()
        version_label = QLabel(f"当前版本: {self.app_version}")
        version_info_layout.addWidget(version_label)
        version_info_layout.addStretch()
        
        version_layout.addLayout(version_info_layout)
        
        check_update_btn = QPushButton("检测更新")
        check_update_btn.clicked.connect(self.check_for_updates)
        version_layout.addWidget(check_update_btn)
        
        version_group.setLayout(version_layout)
        main_layout.addWidget(version_group)
        
        # 导出路径设置
        path_group = QGroupBox("导出路径")
        path_layout = QHBoxLayout()
        
        self.path_input = QLineEdit()
        self.path_input.setText(self.export_path)
        self.path_input.setReadOnly(True)
        path_layout.addWidget(self.path_input)
        
        browse_btn = QPushButton("浏览")
        browse_btn.clicked.connect(self.browse_export_path)
        path_layout.addWidget(browse_btn)
        
        path_group.setLayout(path_layout)
        main_layout.addWidget(path_group)
        
        self.setLayout(main_layout)
    
    def browse_export_path(self):
        """浏览导出路径"""
        folder_path = QFileDialog.getExistingDirectory(
            self, "选择导出路径", self.export_path
        )
        
        if folder_path:
            self.export_path = folder_path.replace('\\', '/')
            self.path_input.setText(self.export_path)
    
    def check_for_updates(self):
        """检测更新（从GitHub）"""
        try:
            import urllib.request
            import json
            import ssl
            from datetime import datetime
            
            # GitHub API URL
            github_api_url = "https://api.github.com/repos/corbin-xu/ToolHub/releases/latest"
            
            print("[DEBUG] ========== 开始检测更新（设置窗口） ==========")
            print(f"[DEBUG] 当前版本: {self.app_version}")
            print(f"[DEBUG] API URL: {github_api_url}")
            
            try:
                print("[DEBUG] 正在创建 SSL 上下文...")
                ssl_context = ssl.create_default_context()
                ssl_context.check_hostname = False
                ssl_context.verify_mode = ssl.CERT_NONE
                print("[DEBUG] SSL 上下文创建成功")
                
                print("[DEBUG] 正在创建请求...")
                request = urllib.request.Request(github_api_url)
                request.add_header('User-Agent', 'ToolHub/1.2')
                print("[DEBUG] 请求创建成功")
                
                print("[DEBUG] 正在连接到 GitHub API...")
                with urllib.request.urlopen(request, context=ssl_context, timeout=10) as response:
                    print(f"[DEBUG] 连接成功！状态码: {response.status}")
                    print("[DEBUG] 正在读取响应数据...")
                    response_data = response.read().decode()
                    print(f"[DEBUG] 响应数据长度: {len(response_data)} 字节")
                    
                    data = json.loads(response_data)
                    latest_version = data.get('tag_name', 'unknown').lstrip('v')
                    print(f"[DEBUG] 最新版本: {latest_version}")
                    
                    # 获取发布日期
                    published_at = data.get('published_at', '')
                    if published_at:
                        # 解析日期格式 (YYYY-MM-DDTHH:MM:SSZ)
                        try:
                            release_date = datetime.fromisoformat(published_at.replace('Z', '+00:00'))
                            self.release_date = release_date.strftime('%Y-%m-%d')
                        except:
                            self.release_date = published_at[:10]
                    
                    if latest_version > self.app_version:
                        print(f"[INFO] 发现新版本: {latest_version}")
                        QMessageBox.information(
                            self, 
                            "有新版本", 
                            f"发现新版本: {latest_version}\n当前版本: {self.app_version}\n\n请访问 GitHub 下载最新版本"
                        )
                    else:
                        print(f"[INFO] 已是最新版本")
                        QMessageBox.information(self, "检测完成", "已是最新版本")
                        
            except urllib.error.URLError as e:
                print(f"[ERROR] URLError 异常!")
                print(f"[ERROR] 错误信息: {e}")
                print(f"[ERROR] 错误原因: {e.reason if hasattr(e, 'reason') else '未知'}")
                import traceback
                traceback.print_exc()
                QMessageBox.warning(self, "检测失败", f"无法连接到 GitHub: {str(e)}")
                
            except urllib.error.HTTPError as e:
                print(f"[ERROR] HTTPError 异常!")
                print(f"[ERROR] HTTP 状态码: {e.code}")
                print(f"[ERROR] 错误信息: {e.reason}")
                QMessageBox.warning(self, "检测失败", f"HTTP 错误 {e.code}: {e.reason}")
                
            except json.JSONDecodeError as e:
                print(f"[ERROR] JSON 解析异常!")
                print(f"[ERROR] 错误信息: {e}")
                QMessageBox.warning(self, "检测失败", "响应数据格式错误")
                
            except Exception as e:
                print(f"[ERROR] 未知异常!")
                print(f"[ERROR] 异常类型: {type(e).__name__}")
                print(f"[ERROR] 错误信息: {str(e)}")
                import traceback
                traceback.print_exc()
                QMessageBox.warning(self, "检测失败", f"检测更新失败: {str(e)}")
                
            print("[DEBUG] ========== 检测更新完成 ==========")
        
        except Exception as e:
            print(f"[ERROR] 外层异常: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "检测失败", f"检测更新失败: {str(e)}")
    
    def get_export_path(self):
        """获取导出路径"""
        return self.export_path


if __name__ == "__main__":
    main()
